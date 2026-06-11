"""
PNL FORECAST -- Flask API Backend
Connects to SQL Server (Actual) & StarRocks (Forecast)
Windows-compatible version
"""

import os
import sys
import io
import json
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime, date
from functools import lru_cache

# ─── Windows UTF-8 safety ───────────────────────────────────
# Prevents UnicodeEncodeError when printing Vietnamese/emoji
# characters to Windows console (default cp1252 / cp65001).
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding='utf-8', errors='replace'
        )
        sys.stderr = io.TextIOWrapper(
            sys.stderr.buffer, encoding='utf-8', errors='replace'
        )

from flask import Flask, jsonify, request, send_from_directory, session, redirect, url_for
from flask_cors import CORS
from dotenv import load_dotenv
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from authlib.integrations.flask_client import OAuth
from werkzeug.security import check_password_hash, generate_password_hash

# ─── Load environment ───
load_dotenv()

# ─── App Setup ───
app = Flask(__name__, static_folder='.', static_url_path='')
app.secret_key = os.getenv('FLASK_SECRET_KEY') or 'pnl_forecast_local_session_key'

# 1. Strict CORS
CORS(app, resources={r"/api/*": {"origins": [
    "http://localhost:5050", 
    "http://127.0.0.1:5050", 
    "https://pnl.ggg.com.vn",
    "https://forecast.ggg.com.vn"
]}}, supports_credentials=True)

# 2. Rate Limiting (chống DDoS)
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per minute", "1000 per hour"],
    storage_uri="memory://"
)

# 3. Authentication SSO
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=os.getenv('GOOGLE_CLIENT_ID', 'placeholder_id'),
    client_secret=os.getenv('GOOGLE_CLIENT_SECRET', 'placeholder_secret'),
    access_token_url='https://accounts.google.com/o/oauth2/token',
    access_token_params=None,
    authorize_url='https://accounts.google.com/o/oauth2/auth',
    authorize_params=None,
    api_base_url='https://www.googleapis.com/oauth2/v1/',
    userinfo_endpoint='https://openidconnect.googleapis.com/v1/userinfo',
    client_kwargs={'scope': 'openid email profile'},
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration'
)

USERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'users.json')

def get_users():
    if not os.path.exists(USERS_FILE):
        return []
    with open(USERS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def filter_restaurants_by_rbac(pc_list):
    """Filters a list of PC codes based on current user session role."""
    if not session.get('user'):
        return pc_list
    user = session['user']
    if user.get('role') == 'admin':
        return pc_list
    allowed_pcs = set(user.get('allowed_pcs', []))
    allowed_chains = set(user.get('allowed_chains', []))
    return [pc for pc in pc_list if pc in allowed_pcs or (len(pc) >= 4 and pc[:4] in allowed_chains)]

@app.before_request
def check_auth():
    """Global API Protection: Check session"""
    if request.path.startswith('/api/') and not request.path.startswith('/api/auth/') and not request.path.startswith('/api/health'):
        if request.method == 'OPTIONS':
            return # Let CORS preflight pass
        
        if 'user' not in session:
            return jsonify({'status': 'error', 'message': 'Unauthorized. Please login.', 'code': 'AUTH_REQUIRED'}), 401
        
        auth_mode = os.getenv('AUTH_MODE', 'local').lower()
        if auth_mode == 'google':
            email = session['user'].get('email', '')
            if not email.endswith('@ggg.com.vn'):
                return jsonify({'status': 'error', 'message': 'Forbidden. GGG emails only.', 'code': 'FORBIDDEN'}), 403


def _require_admin():
    """Return a 403 response if the current session user is not admin, else None."""
    user = session.get('user') or {}
    if user.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Chỉ admin mới có quyền thực hiện thao tác này'}), 403
    return None

# ─── Auth Routes ───
@app.route('/api/auth/login')
def auth_login_page():
    # Serve unified login page
    return send_from_directory('.', 'login.html')

@app.route('/api/auth/google')
def auth_google_redirect():
    redirect_uri = url_for('auth_callback', _external=True)
    return google.authorize_redirect(redirect_uri, prompt='select_account')

@app.route('/api/auth/local/login', methods=['POST'])
def auth_local_login():
    data = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')

    users = get_users()
    matched_user = next((u for u in users if u['username'].lower() == username.lower()), None)
    
    if matched_user and check_password_hash(matched_user['password_hash'], password):
        session['user'] = {
            'username': matched_user['username'],
            'email': matched_user.get('email', username),
            'role': matched_user.get('role', 'user'),
            'allowed_chains': matched_user.get('allowed_chains', []),
            'allowed_pcs': matched_user.get('allowed_pcs', [])
        }
        return jsonify({'status': 'ok'})
    else:
        return jsonify({'status': 'error', 'message': 'Sai tên đăng nhập hoặc mật khẩu'}), 401

@app.route('/api/auth/callback')
def auth_callback():
    try:
        token = google.authorize_access_token()
        user = google.parse_id_token(token)
    except Exception as e:
        return f"""
        <!DOCTYPE html>
        <html><head><title>Lỗi xác thực</title></head>
        <body style="font-family: Arial, sans-serif; text-align: center; padding-top: 100px; background-color: #f9fafb;">
            <h2 style="color: #dc2626; margin-bottom: 5px;">Lỗi Đăng Nhập</h2>
            <p style="color: #4b5563;">{str(e)}</p>
            <a href="/api/auth/login" style="display:inline-block; margin-top:20px; padding:10px 20px; background:#2563eb; color:white; font-weight: bold; text-decoration:none; border-radius:6px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">Thử Đăng nhập Lại</a>
        </body></html>
        """, 400

    if user:
        email = user.get('email', '')
        auth_mode = os.getenv('AUTH_MODE', 'local').lower()
        if auth_mode == 'google' and not email.endswith('@ggg.com.vn'):
            return f"""
            <!DOCTYPE html>
            <html><head><title>Từ chối truy cập</title></head>
            <body style="font-family: Arial, sans-serif; text-align: center; padding-top: 100px; background-color: #f9fafb;">
                <h2 style="color: #dc2626; margin-bottom: 5px;">Truy cập bị từ chối</h2>
                <p style="color: #4b5563; font-size: 1.1rem;">Khách vị <strong>{email}</strong> không thuộc tổ chức nội bộ.</p>
                <p style="color: #6b7280; font-size: 0.9rem;">Vui lòng đăng nhập lại bằng email công ty mang đuôi @ggg.com.vn.</p>
                <a href="/api/auth/login" style="display:inline-block; margin-top:20px; padding:10px 20px; background:#2563eb; color:white; font-weight: bold; text-decoration:none; border-radius:6px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);">Chọn Tài khoản khác</a>
            </body></html>
            """, 403
            
        session['user'] = {
            'username': email.split('@')[0],
            'email': email,
            'role': 'admin' if email == 'ndtrunghieu.nus@gmail.com' else 'user',
            'allowed_chains': [],
            'allowed_pcs': []
        }
        return redirect('/')
    return "Đăng nhập thất bại.", 401

@app.route('/api/auth/logout')
def auth_logout():
    session.pop('user', None)
    return redirect('/')

@app.route('/api/auth/me')
def auth_me():
    if 'user' in session:
        return jsonify({'status': 'ok', 'user': session['user']})
    return jsonify({'status': 'unauthorized'}), 401

@app.route('/api/admin/users', methods=['GET', 'POST'])
def admin_users():
    if session.get('user', {}).get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Forbidden'}), 403
        
    users = get_users()
    
    if request.method == 'GET':
        safe_users = []
        for u in users:
            su = dict(u)
            su.pop('password_hash', None)
            safe_users.append(su)
        return jsonify({'status': 'ok', 'users': safe_users})
        
    elif request.method == 'POST':
        data = request.json or {}
        action = data.get('action')
        username = data.get('username', '').strip()
        
        if not username:
            return jsonify({'status': 'error', 'message': 'Username is required'}), 400
            
        if action == 'create' or action == 'update':
            existing = next((u for u in users if u['username'].lower() == username.lower()), None)
            if action == 'create' and existing:
                return jsonify({'status': 'error', 'message': 'Username already exists'}), 400
                
            password = data.get('password')
            role = data.get('role', 'user')
            allowed_chains = data.get('allowed_chains', [])
            allowed_pcs = data.get('allowed_pcs', [])
            
            if not existing:
                if not password:
                    return jsonify({'status': 'error', 'message': 'Password required for new user'}), 400
                new_user = {
                    'username': username,
                    'password_hash': generate_password_hash(password),
                    'role': role,
                    'allowed_chains': allowed_chains,
                    'allowed_pcs': allowed_pcs
                }
                users.append(new_user)
            else:
                if password:
                    existing['password_hash'] = generate_password_hash(password)
                existing['role'] = role
                existing['allowed_chains'] = allowed_chains
                existing['allowed_pcs'] = allowed_pcs
                
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                json.dump(users, f, indent=4)
                
            return jsonify({'status': 'ok'})
            
        elif action == 'delete':
            if username.lower() == 'admin':
                return jsonify({'status': 'error', 'message': 'Cannot delete default admin'}), 400
            users = [u for u in users if u['username'].lower() != username.lower()]
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                json.dump(users, f, indent=4)
            return jsonify({'status': 'ok'})
            
        return jsonify({'status': 'error', 'message': 'Invalid action'}), 400


# ─── Logging: console + rotating file handler (UTF-8) ───
_log_format = '%(asctime)s [%(levelname)s] %(message)s'
logging.basicConfig(level=logging.INFO, format=_log_format)
logger = logging.getLogger(__name__)

_log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'flask.log')
try:
    _file_handler = RotatingFileHandler(
        _log_file, maxBytes=5 * 1024 * 1024, backupCount=3, encoding='utf-8'
    )
    _file_handler.setFormatter(logging.Formatter(_log_format))
    _file_handler.setLevel(logging.INFO)
    logging.getLogger().addHandler(_file_handler)
except Exception as _fh_err:
    logger.warning(f'Could not attach file log handler: {_fh_err}')

# ═══════════════════════════════════════════════════════════════
# DATABASE CONNECTION HELPERS
# ═══════════════════════════════════════════════════════════════

class MssqlCursorWrapper:
    """
    Thin wrapper around pymssql cursor that supports `as_dict` mode.
    pymssql natively supports as_dict on the cursor; this class provides
    a uniform interface used throughout the app.
    """
    def __init__(self, cursor, as_dict=False, paramstyle='format'):
        self._cursor = cursor
        self.as_dict = as_dict
        self.paramstyle = paramstyle

    def execute(self, *args, **kwargs):
        if self.paramstyle == 'qmark' and args:
            query = args[0].replace('%s', '?') if isinstance(args[0], str) else args[0]
            args = (query, *args[1:])
        self._cursor.execute(*args, **kwargs)

    def _row_to_dict(self, row):
        if row is None or not self.as_dict:
            return row
        if isinstance(row, dict):
            return row
        columns = [col[0] for col in (self._cursor.description or [])]
        return {columns[i]: row[i] for i in range(len(columns))}

    def fetchone(self):
        return self._row_to_dict(self._cursor.fetchone())

    def fetchall(self):
        rows = self._cursor.fetchall()
        if not self.as_dict:
            return rows
        return [self._row_to_dict(row) for row in rows]

    def close(self):
        self._cursor.close()


class MssqlConnWrapper:
    """Thin wrapper around pymssql connection for as_dict cursor support."""
    def __init__(self, conn, as_dict=False, driver='pymssql'):
        self._conn = conn
        self._default_as_dict = as_dict
        self.driver = driver

    def cursor(self, as_dict=False):
        use_dict = as_dict or self._default_as_dict
        if self.driver == 'pyodbc':
            return MssqlCursorWrapper(
                self._conn.cursor(),
                as_dict=use_dict,
                paramstyle='qmark'
            )
        # pymssql supports as_dict natively at cursor creation
        return MssqlCursorWrapper(
            self._conn.cursor(as_dict=use_dict),
            as_dict=use_dict
        )

    def close(self):
        self._conn.close()


import threading
_mssql_local = threading.local()


def get_mssql_connection():
    """
    Return a persistent thread-local SQL Server connection via pymssql (cross-platform).
    Works on both Windows (local) and Linux (Render/Docker) without
    requiring ODBC drivers.
    """
    if not hasattr(_mssql_local, 'conn'):
        _mssql_local.conn = None

    host     = os.getenv('MSSQL_HOST',     '192.168.222.13')
    user     = os.getenv('MSSQL_USER',     'misreader')
    db       = os.getenv('MSSQL_DATABASE', '')
    port     = int(os.getenv('MSSQL_PORT', 1433))
    password = os.getenv('MSSQL_PASSWORD')

    def _connect_pyodbc():
        import pyodbc
        drivers = pyodbc.drivers()
        driver = next((d for d in reversed(drivers) if 'SQL Server' in d), None)
        if not driver:
            raise RuntimeError('No SQL Server ODBC driver installed')
        logger.info(f"[DB] Opening new thread-local pyodbc connection to {host}:{port} as {user}...")
        conn_str = (
            f"DRIVER={{{driver}}};"
            f"SERVER={host},{port};"
            f"UID={user};"
            f"PWD={password or ''};"
            "TrustServerCertificate=yes;"
            "Connection Timeout=5;"
        )
        if db:
            conn_str += f"DATABASE={db};"
        raw_conn = pyodbc.connect(conn_str, timeout=30)
        return MssqlConnWrapper(raw_conn, driver='pyodbc')

    def _connect_pymssql():
        import pymssql  # cross-platform: no ODBC driver needed
        logger.info(f"[DB] Opening new thread-local pymssql connection to {host}:{port} as {user}...")
        raw_conn = pymssql.connect(
            server=host,
            port=port,
            user=user,
            password=password,
            database=db,
            login_timeout=10,
            timeout=30,
            charset='UTF-8'
        )
        return MssqlConnWrapper(raw_conn, driver='pymssql')

    def _connect():
        if os.getenv('MSSQL_USE_PYMSSQL', '').lower() not in ('1', 'true', 'yes'):
            try:
                return _connect_pyodbc()
            except Exception as e:
                logger.warning(f"[DB] pyodbc connection failed, falling back to pymssql: {e}")
        return _connect_pymssql()

    # Ping the existing connection before reusing it
    if _mssql_local.conn is not None:
        try:
            cur = _mssql_local.conn.cursor()
            cur.execute("SELECT 1")
            cur.close()
            return _mssql_local.conn
        except Exception:
            logger.warning("[DB] Stale thread-local connection detected -- reconnecting...")
            try:
                _mssql_local.conn.close()
            except Exception:
                pass
            _mssql_local.conn = None

    try:
        _mssql_local.conn = _connect()
        return _mssql_local.conn
    except Exception as e:
        logger.error(f"[DB] Thread-local connection failed to {host}: {e}")
        raise



def get_starrocks_connection():
    """Connect to StarRocks for Forecast data (MySQL protocol)."""
    import pymysql
    return pymysql.connect(
        host=os.getenv('STARROCKS_HOST', '192.168.221.200'),
        port=int(os.getenv('STARROCKS_PORT', 31234)),
        user=os.getenv('STARROCKS_USER', 'mis_admin'),
        password=os.getenv('STARROCKS_PASSWORD'),
        database=os.getenv('STARROCKS_DATABASE', 'datamart_mis_prod'),
        connect_timeout=15,
        read_timeout=30,
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )


def safe_float(val):
    """Safely convert database value to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# ── Simple in-memory cache (TTL = 60 s) ──────────────────────────
import time as _time
_cache: dict = {}
_CACHE_TTL = 60  # seconds

def _cache_get(key):
    entry = _cache.get(key)
    if entry and (_time.time() - entry['ts'] < _CACHE_TTL):
        return entry['val']
    return None

def _cache_set(key, val):
    _cache[key] = {'ts': _time.time(), 'val': val}

def _cache_bust():
    """Clear all cached responses (call after a data refresh)."""
    _cache.clear()


# ═══════════════════════════════════════════════════════════════
# STATIC FILE SERVING
# ═══════════════════════════════════════════════════════════════

@app.route('/')
def serve_index():
    return send_from_directory('.', 'index.html')


@app.route('/css/<path:filename>')
def serve_css(filename):
    return send_from_directory('css', filename)


@app.route('/js/<path:filename>')
def serve_js(filename):
    return send_from_directory('js', filename)


# ═══════════════════════════════════════════════════════════════
# API 1: ACTUAL DATA (SQL Server)
# Table: DataMart.MIS.FC213_FACT_ACT
# Schema (long/vertical format):
#   pc               - restaurant code
#   indicator_code   - line item code (hierarchical: L1/L2/L3)
#   actual_numerator - the numeric value
#   datakey          - period in YYYYMM format (e.g. 202604)
# ═══════════════════════════════════════════════════════════════

ACTUAL_TABLE = "DataMart.MIS.FC213_FACT_ACT"
DAILY_SALES_TABLE = "DataMart.MIS.DAILY_SALES"


def _month_range_for_datekey(datekey):
    dk = int(datekey)
    year, month = dk // 100, dk % 100
    start = date(year, month, 1)
    end = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return start.isoformat(), end.isoformat()


def _fetch_daily_sales_tc(conn, datekey, pc_raw=None, chain=None):
    """
    Fetch TC (GuestsCount) from DAILY_SALES for a given period.
    Considers both single/multiple restaurants (pc_raw) or chains (chain).
    """
    cursor = conn.cursor(as_dict=True)
    start_date, end_date = _month_range_for_datekey(datekey)

    query = f"""
        SELECT SUM(CAST(ISNULL(GuestsCount, 0) AS FLOAT)) AS total_tc
        FROM {DAILY_SALES_TABLE}
        WHERE ShiftDate >= %s AND ShiftDate < %s
    """
    params = [start_date, end_date]

    # Two-step lookup for target RestaurantCodes to prevent SQL Server deadlocks
    target_rest_codes = set()

    if pc_raw:
        pc_list = [p.strip() for p in pc_raw.split(',') if p.strip()]
        for p in pc_list:
            if len(p) >= 4:
                target_rest_codes.add(p[-4:])
    elif chain:
        chain_list = [c.strip() for c in chain.split(',') if c.strip()]
        if chain_list:
            chain_conds = " OR ".join(["pc LIKE %s"] * len(chain_list))
            chain_params = [f"{c}%" for c in chain_list]
            cursor.execute(f"SELECT DISTINCT RIGHT(RTRIM(pc), 4) AS rcode FROM {ACTUAL_TABLE} WHERE {chain_conds}", tuple(chain_params))
            for row in cursor.fetchall():
                if row.get('rcode'):
                    target_rest_codes.add(row['rcode'])
                    
    target_list = list(target_rest_codes)
    if (pc_raw or chain) and not target_list:
        # Filter provided but matched zero rest codes -> return 0
        cursor.close()
        return 0.0

    if target_list:
        placeholders = ', '.join(['%s'] * len(target_list))
        query += f" AND RestaurantCode IN ({placeholders})"
        params.extend(target_list)

    cursor.execute(query, tuple(params))
    row = cursor.fetchone()
    cursor.close()

    return safe_float(row['total_tc']) if row else 0.0


def _fetch_daily_sales_tc_by_period(conn, datekeys, pc_raw=None, chain=None):
    """
    Fetch TC (GuestsCount) from DAILY_SALES for multiple periods.
    Considers both single/multiple restaurants (pc_raw) or chains (chain).
    Returns: { datekey: total_tc, ... }
    """
    cursor = conn.cursor(as_dict=True)

    conditions = []
    params = []
    for dk in datekeys:
        dk_int = int(dk)
        conditions.append("(YEAR(ShiftDate) = %s AND MONTH(ShiftDate) = %s)")
        params.extend([dk_int // 100, dk_int % 100])

    if not conditions:
        cursor.close()
        return {}

    query = f"""
        SELECT YEAR(ShiftDate) * 100 + MONTH(ShiftDate) AS dk,
               SUM(CAST(ISNULL(GuestsCount, 0) AS FLOAT)) AS total_tc
        FROM {DAILY_SALES_TABLE}
        WHERE ({' OR '.join(conditions)})
    """

    target_rest_codes = set()
    if pc_raw:
        pc_list = [p.strip() for p in pc_raw.split(',') if p.strip()]
        for p in pc_list:
            if len(p) >= 4:
                target_rest_codes.add(p[-4:])
    elif chain:
        chain_list = [c.strip() for c in chain.split(',') if c.strip()]
        if chain_list:
            chain_conds = " OR ".join(["pc LIKE %s"] * len(chain_list))
            chain_params = [f"{c}%" for c in chain_list]
            cursor.execute(f"SELECT DISTINCT RIGHT(RTRIM(pc), 4) AS rcode FROM {ACTUAL_TABLE} WHERE {chain_conds}", tuple(chain_params))
            for row in cursor.fetchall():
                if row.get('rcode'):
                    target_rest_codes.add(row['rcode'])

    target_list = list(target_rest_codes)
    if (pc_raw or chain) and not target_list:
        cursor.close()
        return {}

    if target_list:
        placeholders = ', '.join(['%s'] * len(target_list))
        query += f" AND RestaurantCode IN ({placeholders})"
        params.extend(target_list)

    query += " GROUP BY YEAR(ShiftDate) * 100 + MONTH(ShiftDate)"

    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    cursor.close()

    return {int(r['dk']): safe_float(r['total_tc']) for r in rows}


def _fetch_daily_sales_tc_by_pc_period(conn, datekeys, pcs):
    """Fetch GuestsCount by restaurant and month for bottom-up forecasts."""
    cursor = conn.cursor(as_dict=True)
    conditions = []
    params = []
    for dk in datekeys:
        start_date, end_date = _month_range_for_datekey(dk)
        conditions.append("(ShiftDate >= %s AND ShiftDate < %s)")
        params.extend([start_date, end_date])

    rcode_to_pc = {}
    for pc in pcs or []:
        pc_str = str(pc).strip()
        if len(pc_str) >= 4:
            rcode_to_pc[pc_str[-4:]] = pc_str

    if not conditions or not rcode_to_pc:
        cursor.close()
        return {}

    placeholders = ', '.join(['%s'] * len(rcode_to_pc))
    query = f"""
        SELECT RestaurantCode,
               YEAR(ShiftDate) * 100 + MONTH(ShiftDate) AS dk,
               SUM(CAST(ISNULL(GuestsCount, 0) AS FLOAT)) AS total_tc
        FROM {DAILY_SALES_TABLE}
        WHERE ({' OR '.join(conditions)})
          AND RestaurantCode IN ({placeholders})
        GROUP BY RestaurantCode, YEAR(ShiftDate) * 100 + MONTH(ShiftDate)
    """
    params.extend(rcode_to_pc.keys())
    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    cursor.close()

    out = {}
    for row in rows:
        pc = rcode_to_pc.get(str(row.get('RestaurantCode') or '').strip())
        if pc:
            out.setdefault(pc, {})[int(row['dk'])] = safe_float(row['total_tc'])
    return out


def _pc_filter_clause(pc_raw, params):
    """Return (sql_clause, updated_params) for comma-separated pc list."""
    if not pc_raw:
        return "", params
    pc_list = [p.strip() for p in pc_raw.split(',') if p.strip()]
    if len(pc_list) == 1:
        return " AND pc = %s", params + [pc_list[0]]
    placeholders = ', '.join(['%s'] * len(pc_list))
    return f" AND pc IN ({placeholders})", params + pc_list


@app.route('/api/actual', methods=['GET'])
def get_actual_data():
    """
    Fetch actual PnL rows from FC213_FACT_ACT.

    Query params:
        datakey  : (optional) Period YYYYMM, e.g. 202604
        pc       : (optional) Comma-separated restaurant codes
        from     : (optional) Start period YYYYMM
        to       : (optional) End period YYYYMM
        indicator: (optional) Filter by indicator_code
    Returns rows: [{ pc, indicator_code, actual_numerator, datakey }, ...]
    """
    try:
        datakey   = request.args.get('datakey') or request.args.get('datekey')
        pc_raw    = request.args.get('pc')
        date_from = request.args.get('from')
        date_to   = request.args.get('to')
        indicator = request.args.get('indicator')

        conn   = get_mssql_connection()
        cursor = conn.cursor(as_dict=True)

        query  = f"SELECT pc, indicator_code, actual_numerator, datekey FROM {ACTUAL_TABLE} WHERE 1=1"
        params = []

        if datakey:
            query += " AND datekey = %s"
            params.append(int(datakey))
        if date_from:
            query += " AND datekey >= %s"
            params.append(int(date_from))
        if date_to:
            query += " AND datekey <= %s"
            params.append(int(date_to))
        if indicator:
            query += " AND indicator_code = %s"
            params.append(indicator)

        pc_clause, params = _pc_filter_clause(pc_raw, params)
        query += pc_clause

        # Default: last 6 months if no period filter
        if not datakey and not date_from and not date_to:
            now = datetime.now()
            six_ago_month = now.month - 6
            six_ago_year  = now.year
            if six_ago_month <= 0:
                six_ago_month += 12
                six_ago_year  -= 1
            query += " AND datekey >= %s"
            params.append(six_ago_year * 100 + six_ago_month)

        query += " ORDER BY datekey DESC, pc, indicator_code"

        logger.info(f"[ACTUAL] query params: {params}")
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()

        result = []
        for row in rows:
            result.append({
                'pc':               str(row.get('pc') or ''),
                'indicator_code':   str(row.get('indicator_code') or ''),
                'actual_numerator': safe_float(row.get('actual_numerator')),
                'datakey':          int(row['datekey']) if row.get('datekey') else None
            })

        logger.info(f"[ACTUAL] returned {len(result)} rows")
        return jsonify({
            'status': 'ok',
            'source': 'mssql',
            'table':  ACTUAL_TABLE,
            'count':  len(result),
            'data':   result
        })

    except Exception as e:
        logger.error(f"[ACTUAL] Error: {str(e)}")
        return jsonify({'status': 'error', 'source': 'mssql', 'message': str(e)}), 500


@app.route('/api/actual/columns', methods=['GET'])
def get_actual_columns():
    """Return distinct indicator_codes available in FC213_FACT_ACT."""
    try:
        conn   = get_mssql_connection()
        cursor = conn.cursor(as_dict=True)
        cursor.execute(f"""
            SELECT DISTINCT indicator_code
            FROM {ACTUAL_TABLE}
            WHERE indicator_code IS NOT NULL AND indicator_code != ''
            ORDER BY indicator_code
        """)
        rows = cursor.fetchall()
        cursor.close()

        columns = [{'code': r['indicator_code'], 'type': 'line_item'} for r in rows]
        return jsonify({'status': 'ok', 'columns': columns, 'total': len(columns)})

    except Exception as e:
        logger.error(f"[ACTUAL/COLUMNS] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/actual/restaurants', methods=['GET'])
def get_restaurants():
    """Get distinct list of ACTIVE restaurant codes (pc) from Open_Close.xlsx."""
    try:
        master_list = _read_master_excel()
        active_pcs = [r['pc'] for r in master_list if r['status'] == 'ACTIVE']
        filtered_restaurants = filter_restaurants_by_rbac(active_pcs)

        return jsonify({
            'status':      'ok',
            'restaurants': sorted(filtered_restaurants),
            'count':       len(filtered_restaurants)
        })

    except Exception as e:
        logger.error(f"[RESTAURANTS] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/actual/periods', methods=['GET'])
def get_periods():
    """Get distinct list of available reporting periods from FC213_FACT_ACT."""
    try:
        conn   = get_mssql_connection()
        cursor = conn.cursor(as_dict=True)
        cursor.execute(f"""
            SELECT DISTINCT datekey
            FROM {ACTUAL_TABLE}
            WHERE datekey IS NOT NULL
            ORDER BY datekey DESC
        """)
        rows = cursor.fetchall()
        cursor.close()

        periods = []
        for r in rows:
            dk    = int(r['datekey'])
            year  = dk // 100
            month = dk % 100
            periods.append({
                'datekey': dk,
                'year':    year,
                'month':   month,
                'label':   f"Tháng {month}/{year}"
            })

        return jsonify({'status': 'ok', 'periods': periods, 'count': len(periods)})

    except Exception as e:
        logger.error(f"[PERIODS] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/actual/datekeys', methods=['GET'])
def get_actual_datekeys():
    """Return distinct datakeys from FC213_FACT_ACT, newest first."""
    try:
        conn   = get_mssql_connection()
        cursor = conn.cursor(as_dict=True)
        cursor.execute(f"""
            SELECT DISTINCT TOP 24 datekey
            FROM {ACTUAL_TABLE}
            WHERE datekey IS NOT NULL
            ORDER BY datekey DESC
        """)
        rows = cursor.fetchall()
        cursor.close()

        datakeys = [int(r['datekey']) for r in rows]
        return jsonify({'status': 'ok', 'datekeys': datakeys})

    except Exception as e:
        logger.error(f"[DATEKEYS] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/actual/summary', methods=['GET'])
def get_actual_summary():
    """
    Aggregate actual data for a period → returns dict keyed by indicator_code.
    Long-format: SUM(actual_numerator) GROUP BY indicator_code.

    Query params:
        datakey / datekey : (optional) Period YYYYMM. Auto-uses latest if omitted.
        pc                : (optional) Comma-separated pc codes.
    Returns:
        { "status": "ok", "datekey": 202604, "data": { "DT01": 123456, ... } }
    """
    try:
        datakey = request.args.get('datakey') or request.args.get('datekey')
        pc_raw  = request.args.get('pc')
        chain   = request.args.get('chain')

        # ── Cache hit ──
        cache_key = f"summary:{datakey or 'latest'}:{pc_raw or chain or 'ALL'}"
        cached = _cache_get(cache_key)
        if cached:
            logger.info(f"[ACTUAL/SUMMARY] Cache hit → {cache_key}")
            return jsonify(cached)

        conn   = get_mssql_connection()
        cursor = conn.cursor(as_dict=True)

        # Auto-detect latest datekey if not supplied
        if not datakey:
            cursor.execute(f"""
                SELECT TOP 1 datekey
                FROM {ACTUAL_TABLE}
                WHERE datekey IS NOT NULL
                ORDER BY datekey DESC
            """)
            dk_row = cursor.fetchone()
            if not dk_row:
                cursor.close()
                return jsonify({'status': 'ok', 'data': {}, 'message': 'No data in table'})
            datakey = str(dk_row['datekey'])

        # Build aggregation query: SUM actual_numerator per indicator_code
        query  = f"""
            SELECT indicator_code,
                   SUM(CAST(ISNULL(actual_numerator, 0) AS FLOAT)) AS total
            FROM {ACTUAL_TABLE}
            WHERE datekey = %s
              AND indicator_code IS NOT NULL
              AND indicator_code != ''
        """
        params = [int(datakey)]

        pc_clause, params = _pc_filter_clause(pc_raw, params)
        query += pc_clause
        if not pc_raw and chain:
            chain_list = [c.strip() for c in chain.split(',') if c.strip()]
            if len(chain_list) == 1:
                query += " AND LEFT(pc, 4) = %s"
                params.append(chain_list[0])
            elif len(chain_list) > 1:
                query += f" AND LEFT(pc, 4) IN ({','.join(['%s'] * len(chain_list))})"
                params.extend(chain_list)
        query += " GROUP BY indicator_code"

        logger.info(f"[ACTUAL/SUMMARY] datekey={datakey} pc={pc_raw} chain={chain}")
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()

        if not rows:
            return jsonify({
                'status':  'ok',
                'data':    {},
                'datekey': int(datakey),
                'message': 'No rows for this filter'
            })

        # Build {indicator_code: value} dict
        result = {}
        for row in rows:
            code = str(row['indicator_code']).strip()
            val  = safe_float(row['total'])
            result[code] = val

        # ── Merge TC from DAILY_SALES (GuestsCount) ──
        try:
            tc_val = _fetch_daily_sales_tc(conn, datakey, pc_raw, chain if not pc_raw else None)
            if tc_val > 0:
                result['TC'] = tc_val
                # Compute TA = DT01 / TC (average revenue per guest)
                dt01 = result.get('DT01', 0)
                if tc_val > 0:
                    result['TA'] = round(dt01 / tc_val, 2)
                else:
                    result['TA'] = 0
                logger.info(f"[ACTUAL/SUMMARY] TC={tc_val} from DAILY_SALES, TA={result.get('TA', 0)}")
        except Exception as e:
            logger.warning(f"[ACTUAL/SUMMARY] Could not fetch TC from DAILY_SALES: {e}")

        payload = {
            'status':     'ok',
            'datekey':    int(datakey),
            'restaurant': pc_raw or chain or 'ALL',
            'data':       result
        }
        _cache_set(cache_key, payload)
        return jsonify(payload)


    except Exception as e:
        logger.error(f"[ACTUAL/SUMMARY] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500





# ═══════════════════════════════════════════════════════════════
# API 2: FORECAST DATA (StarRocks)
# Table: DataMart.MIS.v_dim_manual_fc_operation_report_forecast_pnl
# ═══════════════════════════════════════════════════════════════

@app.route('/api/forecast', methods=['GET'])
def get_forecast_data():
    """
    Fetch forecast PnL data from StarRocks.
    
    Query params:
        datekey  : (optional) Filter by specific period
        pc       : (optional) Filter by restaurant code
        from     : (optional) Start period
        to       : (optional) End period
        limit    : (optional) Max rows (default 1000)
    """
    try:
        datekey = request.args.get('datekey')
        pc = request.args.get('pc')
        date_from = request.args.get('from')
        date_to = request.args.get('to')
        limit = request.args.get('limit', 1000, type=int)

        conn = get_starrocks_connection()
        cursor = conn.cursor()

        query = "SELECT * FROM DataMart.MIS.v_dim_manual_fc_operation_report_forecast_pnl WHERE 1=1"
        params = []

        if datekey:
            query += " AND datekey = %s"
            params.append(int(datekey))
        if pc:
            query += " AND pc = %s"
            params.append(pc)
        if date_from:
            query += " AND datekey >= %s"
            params.append(int(date_from))
        if date_to:
            query += " AND datekey <= %s"
            params.append(int(date_to))

        query += f" ORDER BY datekey DESC LIMIT {limit}"

        logger.info(f"[FORECAST] Executing query with params: {params}")
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()

        # DictCursor already returns dicts
        result = []
        for row in rows:
            clean_row = {}
            for key, val in row.items():
                if val is None:
                    clean_row[key] = None
                elif isinstance(val, (int, float)):
                    clean_row[key] = val
                else:
                    try:
                        clean_row[key] = float(val)
                    except (ValueError, TypeError):
                        clean_row[key] = str(val)
            result.append(clean_row)

        cursor.close()

        logger.info(f"[FORECAST] Returned {len(result)} rows")
        return jsonify({
            'status': 'ok',
            'source': 'starrocks',
            'table': 'DataMart.MIS.v_dim_manual_fc_operation_report_forecast_pnl',
            'count': len(result),
            'data': result
        })

    except Exception as e:
        logger.error(f"[FORECAST] Error: {str(e)}")
        return jsonify({
            'status': 'error',
            'source': 'starrocks',
            'message': str(e)
        }), 500


@app.route('/api/forecast/columns', methods=['GET'])
def get_forecast_columns():
    """Get column names from the Forecast table."""
    try:
        conn = get_starrocks_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT * FROM DataMart.MIS.v_dim_manual_fc_operation_report_forecast_pnl LIMIT 1
        """)
        row = cursor.fetchone()

        columns = []
        if row:
            for key in row.keys():
                col_lower = key.lower()
                if col_lower not in ('datekey', 'pc', 'tc'):
                    columns.append({'code': key, 'type': 'line_item'})
                else:
                    columns.append({'code': key, 'type': 'dimension'})

        cursor.close()

        return jsonify({
            'status': 'ok',
            'columns': columns,
            'total': len(columns)
        })

    except Exception as e:
        logger.error(f"[FORECAST/COLUMNS] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════
# API 3: COMBINED — Actual vs Forecast for Dashboard
# ═══════════════════════════════════════════════════════════════

@app.route('/api/pnl', methods=['GET'])
@limiter.limit("20 per minute")
def get_pnl_combined():
    """
    Combined endpoint: fetch both Actual + Forecast for a given period/restaurant.
    Merges data into a unified PnL structure keyed by indicator_code.

    Query params:
        datakey / datekey : (optional) Period YYYYMM. Uses latest if omitted.
        pc                : (optional) Comma-separated restaurant codes
    """
    try:
        datakey = request.args.get('datakey') or request.args.get('datekey')
        pc_raw  = request.args.get('pc')

        actual_data   = {}
        forecast_data = {}
        errors        = []

        # ── Fetch Actual from MSSQL (long-format) ──
        try:
            conn_ms = get_mssql_connection()
            cur_ms  = conn_ms.cursor(as_dict=True)

            # Auto-detect latest datakey
            if not datakey:
                cur_ms.execute(f"SELECT TOP 1 datekey FROM {ACTUAL_TABLE} WHERE datekey IS NOT NULL ORDER BY datekey DESC")
                dk_row = cur_ms.fetchone()
                datakey = str(dk_row['datekey']) if dk_row else None

            if datakey:
                q      = f"""
                    SELECT indicator_code,
                           SUM(CAST(ISNULL(actual_numerator,0) AS FLOAT)) AS total
                    FROM {ACTUAL_TABLE}
                    WHERE datekey = %s
                      AND indicator_code IS NOT NULL AND indicator_code != ''
                """
                params = [int(datakey)]
                pc_clause, params = _pc_filter_clause(pc_raw, params)
                q += pc_clause + " GROUP BY indicator_code"

                cur_ms.execute(q, tuple(params))
                for row in cur_ms.fetchall():
                    actual_data[str(row['indicator_code']).strip()] = safe_float(row['total'])

            cur_ms.close()
            # NOTE: Do NOT close conn_ms here — it is the module-level
            # singleton managed by get_mssql_connection(). Closing it
            # would break subsequent requests that reuse the same socket.
        except Exception as e:
            errors.append(f"MSSQL: {str(e)}")
            logger.error(f"[PNL/ACTUAL] {str(e)}")

        # ── Fetch Forecast from StarRocks (wide-format) ──
        try:
            conn_sr = get_starrocks_connection()
            cur_sr  = conn_sr.cursor()

            q = "SELECT * FROM DataMart.MIS.v_dim_manual_fc_operation_report_forecast_pnl WHERE datekey = %s"
            p = [int(datakey)] if datakey else []
            if not datakey:
                raise ValueError("No datakey available")
            if pc_raw:
                pc_list = [x.strip() for x in pc_raw.split(',') if x.strip()]
                if len(pc_list) == 1:
                    q += " AND pc = %s"; p.append(pc_list[0])
                elif len(pc_list) > 1:
                    q += f" AND pc IN ({','.join(['%s']*len(pc_list))})"; p.extend(pc_list)

            cur_sr.execute(q, tuple(p))
            rows = cur_sr.fetchall()
            if rows:
                for row in rows:
                    for key, val in row.items():
                        if key.lower() not in ('datekey', 'pc'):
                            forecast_data[key] = forecast_data.get(key, 0) + safe_float(val)

            cur_sr.close()
            conn_sr.close()
        except Exception as e:
            errors.append(f"StarRocks: {str(e)}")
            logger.error(f"[PNL/FORECAST] {str(e)}")

        # ── Build unified line items ──
        all_codes  = set(list(actual_data.keys()) + list(forecast_data.keys()))
        line_items = []
        for code in sorted(all_codes):
            act         = actual_data.get(code, 0)
            fct         = forecast_data.get(code, 0)
            variance    = act - fct
            variance_pct = (variance / abs(fct) * 100) if fct != 0 else 0

            line_items.append({
                'code':          code,
                'actual':        round(act, 2),
                'forecast':      round(fct, 2),
                'variance':      round(variance, 2),
                'variance_pct':  round(variance_pct, 2)
            })

        return jsonify({
            'status':     'ok',
            'datakey':    int(datakey) if datakey else None,
            'restaurant': pc_raw or 'ALL',
            'line_items': line_items,
            'errors':     errors if errors else None
        })

    except Exception as e:
        logger.error(f"[PNL] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500



# ═══════════════════════════════════════════════════════════════
# API 4: CHAIN / RESTAURANT HIERARCHY
# Chain = first 4 chars of pc code
# ═══════════════════════════════════════════════════════════════

@app.route('/api/chains', methods=['GET'])
def get_chains():
    """
    Get chain → restaurant hierarchy.
    Chains are identified by the first 4 characters of pc.
    Source: Open_Close.xlsx (Master list, ACTIVE only)
    Returns: { chains: [ { chain: '10GG', name: 'Gogi', restaurants: ['10GG4102', ...] } ] }
    """
    try:
        master_list = _read_master_excel()
        active_master = [r for r in master_list if r.get('status') == 'ACTIVE']
        filtered_set = set(filter_restaurants_by_rbac([r['pc'] for r in active_master]))

        chain_map = {}
        for r in active_master:
            pc = r['pc']
            if pc not in filtered_set or len(pc) < 4:
                continue
            prefix = pc[:4]
            chain_name = r.get('chain_name') or r.get('br') or r.get('brand') or prefix
            group_key = chain_name.upper()
            if group_key not in chain_map:
                chain_map[group_key] = {
                    'chain_name': chain_name,
                    'prefixes': set(),
                    'restaurants': []
                }
            chain_map[group_key]['prefixes'].add(prefix)
            chain_map[group_key]['restaurants'].append({
                'code':   pc,
                'name':   r.get('store', ''),
                'region': r.get('area', ''),
                'area':   r.get('area', ''),
                'brand':  r.get('brand', ''),
                'chain_name': chain_name,
                'status': r.get('status', '')
            })

        chains = []
        for group_key in sorted(chain_map.keys()):
            prefixes = sorted(chain_map[group_key]['prefixes'])
            restaurants = sorted(chain_map[group_key]['restaurants'], key=lambda x: x['code'])
            chains.append({
                'chain':       ','.join(prefixes),
                'name':        chain_map[group_key]['chain_name'],
                'chain_name':  chain_map[group_key]['chain_name'],
                'prefixes':    prefixes,
                'count':       len(restaurants),
                'restaurants': restaurants
            })

        logger.info(f"[CHAINS] Loaded {len(chains)} ACTIVE chains from Open_Close.xlsx after RBAC")

        return jsonify({
            'status':             'ok',
            'chains':             chains,
            'total_chains':       len(chains),
            'total_restaurants':  sum(c['count'] for c in chains)
        })

    except Exception as e:
        logger.error(f"[CHAINS] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500



# ═══════════════════════════════════════════════════════════════
# API 5: MULTI-PERIOD ACTUAL TREND
# For building historical base data for forecast
# ═══════════════════════════════════════════════════════════════

@app.route('/api/actual/trend', methods=['GET'])
def get_actual_trend():
    """
    Get aggregated actual data across multiple periods.
    Long-format: aggregates SUM(actual_numerator) per (datakey, indicator_code).
    Returns pivoted result: [ { datakey, DT01: val, CP01: val, ... }, ... ]

    Query params:
        months : (optional) Number of months to look back (default 12)
        pc     : (optional) Comma-separated restaurant codes
        chain  : (optional) Chain prefix (first 4 chars)
    """
    try:
        months = request.args.get('months', 12, type=int)
        pc_raw = request.args.get('pc')
        chain  = request.args.get('chain')

        conn   = get_mssql_connection()
        cursor = conn.cursor(as_dict=True)

        # Build period list
        now = datetime.now()
        periods = []
        for i in range(months):
            m = now.month - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            periods.append(y * 100 + m)

        where_parts = [f"datekey IN ({','.join(['%s'] * len(periods))})"]
        params = list(periods)

        if pc_raw:
            pc_clause, params = _pc_filter_clause(pc_raw, params)
            where_parts.append(pc_clause.lstrip(" AND "))
        elif chain:
            where_parts.append("LEFT(pc, 4) = %s")
            params.append(chain)

        where_sql = " AND ".join(where_parts)

        query = f"""
            SELECT datekey, indicator_code,
                   SUM(CAST(ISNULL(actual_numerator, 0) AS FLOAT)) AS total
            FROM {ACTUAL_TABLE}
            WHERE {where_sql}
              AND indicator_code IS NOT NULL AND indicator_code != ''
            GROUP BY datekey, indicator_code
            ORDER BY datekey ASC, indicator_code
        """

        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        cursor.close()

        # Pivot into { datakey: { indicator_code: value } }
        pivoted = {}
        for row in rows:
            dk   = int(row['datekey'])
            code = str(row['indicator_code']).strip()
            val  = safe_float(row['total'])
            if dk not in pivoted:
                pivoted[dk] = {'datekey': dk}
            pivoted[dk][code] = val

        result = sorted(pivoted.values(), key=lambda x: x['datekey'])

        return jsonify({
            'status':          'ok',
            'months_requested': months,
            'filter':          pc_raw or chain or 'ALL',
            'periods_found':   len(result),
            'data':            result
        })

    except Exception as e:
        logger.error(f"[ACTUAL/TREND] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500



# ═══════════════════════════════════════════════════════════════
# API 6: FORECAST ENGINE — Compute projections
# ═══════════════════════════════════════════════════════════════

@app.route('/api/forecast/compute', methods=['POST'])
@limiter.limit("15 per minute")
def compute_forecast():
    """
    Compute forecast projections based on actual historical data.

    POST body (JSON):
        horizon   : 1 | 3 | 6 | 12 (months to forecast)
        method    : 'historical' | 'fixed' | 'anchor' | 'percent_revenue'
        pc        : (optional) single restaurant
        chain     : (optional) chain prefix
        params    : { growth_rate, fixed_values, anchor_period, ... }
    """
    # DEPRECATED: this endpoint forecasts on chain/company-aggregated actuals
    # (top-down), which violates the store-first policy, and it accepted
    # arbitrary item_configs from the request body. Use /api/forecast/compute-v2
    # (bottom-up per store, locked formulas) instead.
    return jsonify({
        'status': 'error',
        'message': 'Endpoint deprecated — use /api/forecast/compute-v2 (bottom-up per store)'
    }), 410
    try:
        body = request.get_json() or {}
        horizon = body.get('horizon', 1)
        method = body.get('method', 'historical')
        pc = body.get('pc')
        chain = body.get('chain')
        method_params = body.get('params', {})

        # ── Step 1: Fetch historical actual data from FC213_FACT_ACT (long-format) ──
        conn = get_mssql_connection()
        cursor = conn.cursor(as_dict=True)

        now = datetime.now()
        periods = []
        for i in range(12):
            m = now.month - i
            y = now.year
            while m <= 0:
                m += 12
                y -= 1
            periods.append(y * 100 + m)

        where_parts = [f"datekey IN ({','.join(['%s'] * len(periods))})"]
        params_sql  = list(periods)

        if pc:
            pc_clause, params_sql = _pc_filter_clause(pc if isinstance(pc, str) else ','.join(pc), params_sql)
            where_parts.append(pc_clause.lstrip(" AND "))
        elif chain:
            # Handle comma-separated multiple chains (e.g. "0000,01HP,03HN")
            chain_list = [c.strip() for c in (chain if isinstance(chain, str) else ','.join(chain)).split(',') if c.strip()]
            if len(chain_list) == 1:
                where_parts.append("LEFT(pc, 4) = %s")
                params_sql.append(chain_list[0])
            else:
                placeholders = ', '.join(['%s'] * len(chain_list))
                where_parts.append(f"LEFT(pc, 4) IN ({placeholders})")
                params_sql.extend(chain_list)

        where_parts.append("indicator_code IS NOT NULL AND indicator_code != ''")

        query = f"""
            SELECT datekey, indicator_code,
                   SUM(CAST(ISNULL(actual_numerator, 0) AS FLOAT)) AS total
            FROM {ACTUAL_TABLE}
            WHERE {' AND '.join(where_parts)}
            GROUP BY datekey, indicator_code
            ORDER BY datekey ASC, indicator_code
        """

        cursor.execute(query, tuple(params_sql))
        raw_rows = cursor.fetchall()
        cursor.close()

        # Pivot long-format into { datakey: { indicator_code: value } }
        pivoted_hist = {}
        line_item_codes_set = set()
        for row in raw_rows:
            dk   = int(row['datekey'])
            code = str(row['indicator_code']).strip()
            val  = safe_float(row['total'])
            if dk not in pivoted_hist:
                pivoted_hist[dk] = {}
            pivoted_hist[dk][code] = val
            line_item_codes_set.add(code)

        hist_rows        = [dict(datekey=dk, **vals) for dk, vals in sorted(pivoted_hist.items())]
        line_item_codes  = sorted(line_item_codes_set)

        # ── Merge TC from DAILY_SALES into historical data ──
        try:
            hist_datekeys = sorted(pivoted_hist.keys())
            tc_by_period = _fetch_daily_sales_tc_by_period(
                conn, hist_datekeys, pc_raw=pc, chain=chain
            )
            for hr in hist_rows:
                dk = int(hr['datekey'])
                tc_val = tc_by_period.get(dk, 0)
                if tc_val > 0:
                    hr['TC'] = tc_val
                    dt01 = safe_float(hr.get('DT01', 0))
                    hr['TA'] = round(dt01 / tc_val, 2) if tc_val > 0 else 0
                    if 'TC' not in line_item_codes_set:
                        line_item_codes_set.add('TC')
                    if 'TA' not in line_item_codes_set:
                        line_item_codes_set.add('TA')
            line_item_codes = sorted(line_item_codes_set)
            logger.info(f"[FORECAST/COMPUTE] Merged TC from DAILY_SALES for {len(tc_by_period)} periods")
        except Exception as e:
            logger.warning(f"[FORECAST/COMPUTE] Could not merge TC from DAILY_SALES: {e}")

        if not hist_rows:
            return jsonify({'status': 'ok', 'message': 'No historical data found', 'projections': []})

        # ── Step 2: Build historical series per line item ──
        historical   = {}
        hist_periods = []
        for row in hist_rows:
            dk = int(row['datekey'])
            hist_periods.append(dk)
            for code in line_item_codes:
                if code not in historical:
                    historical[code] = []
                historical[code].append({
                    'datekey': dk,
                    'value':   safe_float(row.get(code, 0))
                })

        # ── Step 3: Compute projections ──
        last_period = hist_periods[-1] if hist_periods else now.year * 100 + now.month
        future_periods = []
        lp_year = last_period // 100
        lp_month = last_period % 100
        for i in range(1, horizon + 1):
            fm = lp_month + i
            fy = lp_year
            while fm > 12:
                fm -= 12
                fy += 1
            future_periods.append(fy * 100 + fm)

        # ── Step 2b: Read per-item configs from request ──
        item_configs = body.get('item_configs', {})

        projections = []
        for fp in future_periods:
            period_data = {'datekey': fp}
            idx = future_periods.index(fp) + 1

            for code in line_item_codes:
                series = historical.get(code, [])
                values = [s['value'] for s in series]

                # Resolve per-item method (fallback to global)
                item_cfg = item_configs.get(code, {})
                item_method = item_cfg.get('method', method)

                if item_method == 'historical':
                    growth_rate = (item_cfg.get('growth_rate', method_params.get('growth_rate', 5.0))) / 100.0
                    lookback = item_cfg.get('lookback', method_params.get('lookback', 3))
                    recent = values[-lookback:] if len(values) >= lookback else values
                    avg_base = sum(recent) / len(recent) if recent else 0
                    period_data[code] = round(avg_base * ((1 + growth_rate) ** idx), 2)

                elif item_method == 'fixed':
                    fixed_val = item_cfg.get('fixed_value')
                    if fixed_val is not None:
                        period_data[code] = fixed_val
                    else:
                        fixed_vals = method_params.get('fixed_values', {})
                        period_data[code] = fixed_vals.get(code, values[-1] if values else 0)

                elif item_method == 'anchor':
                    anchor_dk = item_cfg.get('anchor_period', method_params.get('anchor_period'))
                    multiplier = item_cfg.get('multiplier', method_params.get('multiplier', 1.05))
                    buffer_val = item_cfg.get('buffer', method_params.get('buffer', 0))
                    anchor_val = 0
                    
                    if str(anchor_dk).startswith('avg_'):
                        months_to_avg = int(str(anchor_dk).split('_')[1])
                        recent_vals = values[-months_to_avg:] if len(values) >= months_to_avg else values
                        anchor_val = sum(recent_vals) / len(recent_vals) if recent_vals else 0
                    else:
                        anchor_dk = int(anchor_dk) if anchor_dk else (hist_periods[-1] if hist_periods else None)
                        for s in series:
                            if s['datekey'] == anchor_dk:
                                anchor_val = s['value']
                                break
                        if anchor_val == 0 and values:
                            anchor_val = values[-1]
                            
                    period_data[code] = round(anchor_val * (multiplier ** idx) + buffer_val, 2)

                elif item_method == 'percent_revenue':
                    base_dk = item_cfg.get('base_period', method_params.get('base_period',
                        hist_periods[-1] if hist_periods else None))
                    
                    if str(base_dk).startswith('avg_'):
                        months_to_avg = int(str(base_dk).split('_')[1])
                        recent_dks = hist_periods[-months_to_avg:] if len(hist_periods) >= months_to_avg else hist_periods
                        sum_item = 0
                        sum_rev = 0
                        for row_h in hist_rows:
                            if int(row_h['datekey']) in recent_dks:
                                sum_item += safe_float(row_h.get(code, 0))
                                dt01 = safe_float(row_h.get('DT01', 0))
                                dt02 = safe_float(row_h.get('DT02', 0))
                                dt03 = safe_float(row_h.get('DT03', 0))
                                dt04 = safe_float(row_h.get('DT04', 0))
                                sum_rev += (dt01 - dt02 - dt03 - dt04)
                        ratio = sum_item / sum_rev if sum_rev != 0 else 0
                    else:
                        base_dk = int(base_dk) if base_dk else (hist_periods[-1] if hist_periods else None)
                        base_item_val = 0
                        base_revenue = 0
                        for row_h in hist_rows:
                            if int(row_h['datekey']) == base_dk:
                                base_item_val = safe_float(row_h.get(code, 0))
                                dt01 = safe_float(row_h.get('DT01', 0))
                                dt02 = safe_float(row_h.get('DT02', 0))
                                dt03 = safe_float(row_h.get('DT03', 0))
                                dt04 = safe_float(row_h.get('DT04', 0))
                                base_revenue = dt01 - dt02 - dt03 - dt04
                                break
                        ratio = base_item_val / base_revenue if base_revenue != 0 else 0

                    rev_growth = (item_cfg.get('revenue_growth', method_params.get('revenue_growth', 5.0))) / 100.0
                    last_rev = 0
                    if hist_rows:
                        last_row = hist_rows[-1]
                        last_rev = (safe_float(last_row.get('DT01', 0))
                                    - safe_float(last_row.get('DT02', 0))
                                    - safe_float(last_row.get('DT03', 0))
                                    - safe_float(last_row.get('DT04', 0)))
                    proj_rev = last_rev * ((1 + rev_growth) ** idx)
                    period_data[code] = round(proj_rev * ratio, 2)

                else:
                    period_data[code] = values[-1] if values else 0

            projections.append(period_data)

        # ── Step 4: Fetch previous forecast for comparison ──
        prev_forecast = {}
        try:
            conn_sr = get_starrocks_connection()
            cur_sr = conn_sr.cursor()

            prev_dk = hist_periods[-1] if hist_periods else None
            if prev_dk:
                q = "SELECT * FROM DataMart.MIS.v_dim_manual_fc_operation_report_forecast_pnl WHERE datekey = %s"
                params_sr = [prev_dk]
                if pc:
                    pc_list_sr = [x.strip() for x in pc.split(',') if x.strip()]
                    if len(pc_list_sr) == 1:
                        q += " AND pc = %s"
                        params_sr.append(pc_list_sr[0])
                    else:
                        q += f" AND pc IN ({','.join(['%s'] * len(pc_list_sr))})"
                        params_sr.extend(pc_list_sr)
                elif chain:
                    chain_list = [c.strip() for c in chain.split(',') if c.strip()]
                    if len(chain_list) == 1:
                        q += " AND LEFT(pc, 4) = %s"
                        params_sr.append(chain_list[0])
                    else:
                        q += f" AND LEFT(pc, 4) IN ({','.join(['%s'] * len(chain_list))})"
                        params_sr.extend(chain_list)

                cur_sr.execute(q, tuple(params_sr))
                fc_rows = cur_sr.fetchall()
                for row in fc_rows:
                    for key, val in row.items():
                        kl = key.lower()
                        if kl not in ('datekey', 'pc'):
                            prev_forecast[key] = prev_forecast.get(key, 0) + safe_float(val)

            cur_sr.close()
            conn_sr.close()
        except Exception as e:
            logger.warning(f"[FORECAST/COMPUTE] Could not fetch previous forecast: {str(e)}")

        return jsonify({
            'status': 'ok',
            'method': method,
            'horizon': horizon,
            'filter': pc or chain or 'ALL',
            'historical_periods': hist_periods,
            'line_items': line_item_codes,
            'historical': [{k: (int(v) if k == 'datekey' else v) for k, v in row.items()} for row in hist_rows],
            'projections': projections,
            'previous_forecast': prev_forecast if prev_forecast else None
        })

    except Exception as e:
        logger.error(f"[FORECAST/COMPUTE] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════
# HEALTH CHECK & CONNECTION TEST
# ═══════════════════════════════════════════════════════════════

def _resolve_master_pcs(pc_raw=None, chain_raw=None):
    master = _read_master_excel()
    active = [r['pc'] for r in master if r.get('status') == 'ACTIVE']
    allowed = set(filter_restaurants_by_rbac(active))
    if pc_raw:
        requested = [p.strip() for p in str(pc_raw).split(',') if p.strip()]
        return [p for p in requested if p in allowed]
    if chain_raw:
        prefixes = [c.strip() for c in str(chain_raw).split(',') if c.strip()]
        return [p for p in active if p in allowed and any(p.startswith(pref) for pref in prefixes)]
    return [p for p in active if p in allowed]


@app.route('/api/forecast/compute-v2', methods=['POST'])
@limiter.limit("15 per minute")
def compute_forecast_v2():
    """Bottom-up forecast: compute each restaurant first, then roll up."""
    try:
        body = request.get_json() or {}
        horizon = int(body.get('horizon', 1) or 1)
        pc = body.get('pc')
        chain = body.get('chain')
        selected_pcs = _resolve_master_pcs(pc, chain)
        all_scope_pcs = _resolve_master_pcs()
        scope_coverage = (len(selected_pcs) / len(all_scope_pcs)) if all_scope_pcs else 0.0
        # The Excel benchmark sheet is an aggregate target. It is used for
        # reconciliation only; store forecasts must remain bottom-up.
        benchmark_scope_ok = scope_coverage >= 0.90
        should_calibrate_to_excel = bool(body.get('force_calibrate_to_excel', False)) and benchmark_scope_ok
        should_benchmark = bool(body.get('benchmark', True)) and benchmark_scope_ok
        if not selected_pcs:
            return jsonify({'status': 'error', 'message': 'No active restaurants matched this filter'}), 400

        conn = get_mssql_connection()
        cursor = conn.cursor(as_dict=True)
        pc_placeholders = ','.join(['%s'] * len(selected_pcs))

        cursor.execute(f"""
            SELECT DISTINCT TOP 12 datekey
            FROM {ACTUAL_TABLE}
            WHERE pc IN ({pc_placeholders})
              AND datekey IS NOT NULL
            ORDER BY datekey DESC
        """, tuple(selected_pcs))
        period_rows = cursor.fetchall()
        periods = [int(r['datekey']) for r in period_rows if r.get('datekey')]
        if not periods:
            cursor.close()
            return jsonify({'status': 'ok', 'message': 'No historical data found', 'projections': [], 'restaurant_projections': {}})

        period_placeholders = ','.join(['%s'] * len(periods))
        cursor.execute(f"""
            SELECT pc, datekey, indicator_code,
                   SUM(CAST(ISNULL(actual_numerator, 0) AS FLOAT)) AS total
            FROM {ACTUAL_TABLE}
            WHERE datekey IN ({period_placeholders})
              AND pc IN ({pc_placeholders})
              AND indicator_code IS NOT NULL AND indicator_code != ''
            GROUP BY pc, datekey, indicator_code
            ORDER BY pc, datekey
        """, tuple(periods + selected_pcs))
        rows = cursor.fetchall()
        cursor.close()

        hist_by_pc = {}
        codes = set()
        for row in rows:
            p = str(row['pc']).strip()
            dk = int(row['datekey'])
            code = str(row['indicator_code']).strip()
            val = safe_float(row['total'])
            hist_by_pc.setdefault(p, {}).setdefault(dk, {})[code] = val
            codes.add(code)

        hist_periods = sorted({int(r['datekey']) for r in rows})
        if not hist_periods:
            return jsonify({'status': 'ok', 'message': 'No historical data found', 'projections': [], 'restaurant_projections': {}})

        try:
            tc_by_pc_period = _fetch_daily_sales_tc_by_pc_period(conn, hist_periods, selected_pcs)
            for p, by_period in tc_by_pc_period.items():
                for dk, tc_val in by_period.items():
                    if tc_val <= 0:
                        continue
                    hist_by_pc.setdefault(p, {}).setdefault(dk, {})['TC'] = tc_val
                    dt01 = safe_float(hist_by_pc[p][dk].get('DT01', 0))
                    hist_by_pc[p][dk]['TA'] = round(dt01 / tc_val, 2) if tc_val else 0.0
                    codes.add('TC')
                    codes.add('TA')
            logger.info(f"[FORECAST/COMPUTE-V2] Merged TC by restaurant for {len(tc_by_pc_period)} stores")
        except Exception as e:
            logger.warning(f"[FORECAST/COMPUTE-V2] Could not merge TC by restaurant: {e}")

        last_period = hist_periods[-1]
        include_current_month = bool(body.get('include_current_month', True))
        future_periods = []
        y, m = last_period // 100, last_period % 100
        start_offset = 0 if include_current_month else 1
        for i in range(start_offset, start_offset + horizon):
            fm = m + i
            fy = y
            while fm > 12:
                fm -= 12
                fy += 1
            future_periods.append(fy * 100 + fm)

        params = body.get('params') or {}
        # Formula logic is LOCKED: forecast_formulas.json is the single source of
        # truth (per the approved Finance rule file). Runtime overrides sent in the
        # request body are ignored so no client can change forecast logic by hand.
        item_configs = _load_forecast_formula_configs()
        if body.get('item_configs'):
            logger.warning("[FORECAST/COMPUTE-V2] Ignored item_configs override from request body (formula config is locked)")
        # Tiền thuê CP0209 theo hợp đồng của từng profit center (import qua /api/rental/import).
        # Store có dữ liệu import → dùng số hợp đồng; store chưa có → fallback theo config (flat LM).
        rental_costs = _load_rental_costs()
        default_growth = safe_float(params.get('growth_rate', 5.0)) / 100.0
        lookback = int(params.get('lookback', 3) or 3)
        # CP07 (Mgt. bonus = 8% lợi nhuận trước bonus) và CP08 (CIT = 20% SD08)
        # là dòng công thức — luôn tính trong _apply_pnl_formula_fields, không forecast trực tiếp.
        formula_codes = {'TA', 'SD01', 'SD02', 'SD03', 'SD04', 'SD05', 'SD07', 'SD08', 'SD09', 'SD10', 'SD11', 'CP07', 'CP08'}

        restaurant_projections = {}
        for p in selected_pcs:
            store_hist = []
            for dk in hist_periods:
                row = {'datekey': dk}
                row.update(hist_by_pc.get(p, {}).get(dk, {}))
                store_hist.append(row)
            store_proj = []
            # Deterministic order: TC and DT01 are projected first because
            # percent_revenue items depend on the projected DT01 of the same period.
            forecastable = sorted(c for c in codes if c not in formula_codes)
            head_codes = [c for c in ('TC', 'DT01') if c in forecastable]
            tail_codes = [c for c in forecastable if c not in ('TC', 'DT01')]
            for idx, fp in enumerate(future_periods):
                out = {'datekey': fp}
                growth_period = idx if include_current_month else idx + 1

                def _project(code):
                    vals = [safe_float(r.get(code, 0)) for r in store_hist if code in r]
                    recent = vals[-lookback:] if vals else []
                    base = sum(recent) / len(recent) if recent else 0.0
                    cfg = item_configs.get(code, {})
                    method = cfg.get('method', 'historical')

                    if method == 'fixed':
                        return round(safe_float(cfg.get('fixed_value', base)), 2)
                    elif method == 'anchor':
                        try:
                            anchor = int(cfg.get('anchor_period') or 0)
                        except (TypeError, ValueError):
                            anchor = 0
                        anchor_row = next((r for r in store_hist if int(r.get('datekey', 0)) == anchor), None)
                        anchor_val = safe_float(anchor_row.get(code, 0)) if anchor_row else (vals[-1] if vals else 0.0)
                        multiplier = safe_float(cfg.get('multiplier', 1.0))
                        buffer_val = safe_float(cfg.get('buffer', 0))
                        return round(anchor_val * multiplier + buffer_val, 2)
                    elif method == 'percent_revenue':
                        if cfg.get('ratio_percent') is not None:
                            # Tỷ lệ cố định trên doanh thu (vd DT03 = 0.95% bình quân 2025)
                            ratio = safe_float(cfg.get('ratio_percent')) / 100.0
                        else:
                            # %LM: lookback=1 lấy tỷ lệ trên doanh thu của tháng gần nhất
                            pr_lookback = int(cfg.get('lookback', lookback) or lookback)
                            ratio_vals = []
                            for r in store_hist[-pr_lookback:]:
                                rev = safe_float(r.get('DT01', 0))
                                if rev:
                                    ratio_vals.append(safe_float(r.get(code, 0)) / rev)
                            ratio = sum(ratio_vals) / len(ratio_vals) if ratio_vals else 0.0
                        return round(safe_float(out.get('DT01', 0)) * ratio, 2)
                    elif method == 'fixed_variable':
                        # "Variable tháng T + fix tháng T": phần fix giữ nguyên mức tháng gần nhất
                        # của chính store đó; phần variable scale theo doanh thu dự phóng / doanh thu
                        # tháng gần nhất. variable_percent = tỷ trọng biến phí (0-100).
                        var_share = min(max(safe_float(cfg.get('variable_percent', 50)) / 100.0, 0.0), 1.0)
                        lm_row = next((r for r in reversed(store_hist) if code in r), None)
                        lm_val = safe_float(lm_row.get(code, 0)) if lm_row else 0.0
                        lm_rev = safe_float(lm_row.get('DT01', 0)) if lm_row else 0.0
                        proj_rev = safe_float(out.get('DT01', 0))
                        rev_ratio = (proj_rev / lm_rev) if lm_rev > 0 else 1.0
                        return round(lm_val * ((1 - var_share) + var_share * rev_ratio), 2)
                    elif method == 'rolling4w':
                        adj = safe_float(cfg.get('rolling4w_adjustment', default_growth * 100)) / 100.0
                        return round(base * ((1 + adj) ** growth_period), 2)
                    else:
                        growth = safe_float(cfg.get('growth_rate', default_growth * 100)) / 100.0
                        cfg_lookback = int(cfg.get('lookback', lookback) or lookback)
                        cfg_vals = vals[-cfg_lookback:] if vals else []
                        cfg_base = sum(cfg_vals) / len(cfg_vals) if cfg_vals else base
                        return round(cfg_base * ((1 + growth) ** growth_period), 2)

                for code in head_codes:
                    out[code] = _project(code)

                tc = safe_float(out.get('TC', 0))
                if tc and not safe_float(out.get('DT01', 0)):
                    ta_vals = [safe_float(r.get('TA', 0)) for r in store_hist if safe_float(r.get('TA', 0)) > 0]
                    ta_recent = ta_vals[-lookback:] if ta_vals else []
                    ta_base = sum(ta_recent) / len(ta_recent) if ta_recent else 0.0
                    if ta_base:
                        out['DT01'] = round(tc * ta_base, 2)

                for code in tail_codes:
                    out[code] = _project(code)

                # Ghi đè CP0209 bằng tiền thuê hợp đồng đã import cho store này (nếu có).
                # Scale các mã con CP0209xx theo cùng tỷ lệ để tổng con khớp số hợp đồng,
                # vì _apply_pnl_formula_fields sẽ tính lại cha = tổng các con.
                rent = rental_costs.get(p)
                if rent is not None and rent > 0:
                    child_keys = [k for k in out
                                  if isinstance(k, str) and len(k) == 8 and k.startswith('CP0209')
                                  and not k.endswith(('PT', 'PMS'))]
                    child_sum = sum(safe_float(out[k]) for k in child_keys)
                    if child_keys and child_sum:
                        factor = rent / child_sum
                        for k in child_keys:
                            out[k] = round(safe_float(out[k]) * factor, 2)
                    out['CP0209'] = round(rent, 2)

                store_proj.append(_apply_pnl_formula_fields(out))
            restaurant_projections[p] = store_proj

        calibration = None
        if should_calibrate_to_excel:
            calibration = _calibrate_restaurant_projections_to_benchmark(restaurant_projections, formula_codes)

        projections = []
        for idx, fp in enumerate(future_periods):
            roll = {'datekey': fp}
            roll_codes = set(codes)
            for values in restaurant_projections.values():
                roll_codes.update(values[idx].keys())
            for code in roll_codes:
                if code in ('datekey', 'TA'):
                    continue
                roll[code] = round(sum(safe_float(v[idx].get(code, 0)) for v in restaurant_projections.values()), 2)
            roll = _apply_pnl_formula_fields(roll, recompute_parents=False)
            if should_calibrate_to_excel:
                roll = _apply_benchmark_overrides(roll, fp)
            projections.append(roll)

        reconciliation = None
        if should_benchmark and projections:
            target_period = int(body.get('benchmark_period') or projections[0].get('datekey'))
            target_projection = next((p for p in projections if int(p.get('datekey', 0)) == target_period), projections[0])
            tolerance_pct = safe_float(body.get('tolerance_pct', 5.0))
            reconciliation = _reconcile_projection_with_benchmark(target_projection, target_period, tolerance_pct)

        return jsonify({
            'status': 'ok',
            'model': 'bottom_up_store',
            'horizon': horizon,
            'filter': pc or chain or 'ALL',
            'historical_periods': hist_periods,
            'line_items': sorted(codes.union({'TA'})),
            'historical': [],
            'projections': projections,
            'restaurant_projections': restaurant_projections,
            'calibration': calibration,
            'calibration_scope': {
                'selected_restaurants': len(selected_pcs),
                'all_scope_restaurants': len(all_scope_pcs),
                'coverage_pct': round(scope_coverage * 100.0, 2),
                'excel_calibration_applied': should_calibrate_to_excel,
                'reason': (
                    'forced_aggregate_calibration'
                    if should_calibrate_to_excel
                    else 'benchmark_reconciliation_only_store_bottom_up'
                    if benchmark_scope_ok
                    else 'filtered_scope_keeps_store_bottom_up'
                )
            },
            'reconciliation': reconciliation
        })
    except Exception as e:
        logger.error(f"[FORECAST/COMPUTE-V2] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


SAVED_REPORTS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'saved_reports.json')
FORECAST_ARCHIVE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'forecast_period_archive.json')
FORECAST_FORMULAS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'forecast_formulas.json')
RENTAL_COSTS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'rental_costs.json')
FORECAST_BENCHMARK_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'forecast_excel_benchmark.json')


def _load_forecast_formula_configs():
    if os.path.exists(FORECAST_FORMULAS_FILE):
        try:
            with open(FORECAST_FORMULAS_FILE, 'r', encoding='utf-8') as f:
                return (json.load(f) or {}).get('formulas', {})
        except Exception as e:
            logger.warning(f"[FORECAST_FORMULAS] Load failed: {e}")
    return {}


# Forecast logic is locked by default: forecast_formulas.json on the server is the
# single source of truth, maintained per the approved Finance rule file
# (Gop_y_Fin_rule_TC_forecast.xlsx). Set ALLOW_FORMULA_EDIT=true in the environment
# only for a deliberate, temporary unlock.
FORMULA_EDIT_UNLOCKED = os.environ.get('ALLOW_FORMULA_EDIT', 'false').strip().lower() == 'true'


@app.route('/api/forecast/formulas', methods=['GET'])
def get_forecast_formulas():
    meta = {}
    if os.path.exists(FORECAST_FORMULAS_FILE):
        try:
            with open(FORECAST_FORMULAS_FILE, 'r', encoding='utf-8') as f:
                meta = json.load(f) or {}
        except Exception as e:
            logger.warning(f"[FORECAST_FORMULAS] Load failed: {e}")
    role = (session.get('user') or {}).get('role')
    return jsonify({
        'status': 'ok',
        'locked': not FORMULA_EDIT_UNLOCKED,
        'can_edit_variable_split': role in ('admin', 'manager'),
        'updated_at': meta.get('updated_at'),
        'updated_by': meta.get('updated_by'),
        'formulas': meta.get('formulas', {})
    })


@app.route('/api/forecast/formulas/variable-split', methods=['POST'])
def update_variable_split():
    """Carve-out duy nhất khỏi cơ chế khoá công thức: manager (hoặc admin) được
    quyết định tỷ trọng biến phí/định phí (variable_percent) của các item
    fixed_variable. Method và mọi tham số khác vẫn khoá."""
    user = session.get('user') or {}
    if user.get('role') not in ('admin', 'manager'):
        return jsonify({'status': 'error', 'message': 'Chỉ manager hoặc admin được chỉnh tỷ trọng biến phí'}), 403
    body = request.get_json() or {}
    splits = body.get('splits')
    if not isinstance(splits, dict) or not splits:
        return jsonify({'status': 'error', 'message': 'Missing splits payload'}), 400

    meta = {}
    if os.path.exists(FORECAST_FORMULAS_FILE):
        try:
            with open(FORECAST_FORMULAS_FILE, 'r', encoding='utf-8') as f:
                meta = json.load(f) or {}
        except Exception as e:
            logger.error(f"[FORECAST_FORMULAS] Load failed: {e}")
            return jsonify({'status': 'error', 'message': 'Không đọc được file công thức'}), 500
    formulas = meta.get('formulas', {})

    updated, rejected = [], []
    for code, pct in splits.items():
        cfg = formulas.get(str(code))
        if not cfg or cfg.get('method') != 'fixed_variable':
            rejected.append(f"{code}: không phải item Fix+Variable")
            continue
        try:
            val = float(pct)
        except (TypeError, ValueError):
            rejected.append(f"{code}: giá trị không hợp lệ")
            continue
        if not 0 <= val <= 100:
            rejected.append(f"{code}: phải nằm trong 0–100")
            continue
        cfg['variable_percent'] = round(val, 1)
        updated.append(code)

    if not updated:
        return jsonify({'status': 'error', 'message': '; '.join(rejected) or 'Không có item hợp lệ'}), 400

    meta['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    meta['updated_by'] = f"{user.get('username') or user.get('email')} (variable split)"
    with open(FORECAST_FORMULAS_FILE, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    logger.info(f"[FORECAST_FORMULAS] Variable split updated for {updated} by '{meta['updated_by']}'")
    msg = f"Đã cập nhật tỷ trọng biến phí cho {len(updated)} item"
    if rejected:
        msg += f" (bỏ qua: {'; '.join(rejected)})"
    return jsonify({'status': 'ok', 'message': msg, 'updated': updated})


# ─── Rental costs (CP0209) per profit center ───
def _load_rental_costs():
    """Return {pc: monthly_rent_amount} from the imported rental file."""
    if os.path.exists(RENTAL_COSTS_FILE):
        try:
            with open(RENTAL_COSTS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f) or {}
            return {str(pc): safe_float(info.get('amount'))
                    for pc, info in (data.get('costs') or {}).items()}
        except Exception as e:
            logger.warning(f"[RENTAL] Load failed: {e}")
    return {}


@app.route('/api/rental/costs', methods=['GET'])
def get_rental_costs():
    meta = {}
    if os.path.exists(RENTAL_COSTS_FILE):
        try:
            with open(RENTAL_COSTS_FILE, 'r', encoding='utf-8') as f:
                meta = json.load(f) or {}
        except Exception as e:
            logger.warning(f"[RENTAL] Load failed: {e}")
    return jsonify({
        'status': 'ok',
        'updated_at': meta.get('updated_at'),
        'updated_by': meta.get('updated_by'),
        'count': len(meta.get('costs') or {}),
        'costs': meta.get('costs', {})
    })


@app.route('/api/rental/template', methods=['GET'])
def download_rental_template():
    """Excel template: danh sách PC đang ACTIVE + tiền thuê hiện có (nếu đã import)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rental CP0209"

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws.append(["PC Code", "Tên nhà hàng (tham khảo)", "Tiền thuê/tháng (VND)", "Ghi chú"])
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (18, 40, 22, 30)[col_idx - 1]

    current = {}
    if os.path.exists(RENTAL_COSTS_FILE):
        try:
            with open(RENTAL_COSTS_FILE, 'r', encoding='utf-8') as f:
                current = (json.load(f) or {}).get('costs', {})
        except Exception:
            pass
    try:
        master = _read_master_excel()
    except Exception as e:
        logger.warning(f"[RENTAL] Master list unavailable for template: {e}")
        master = []
    for r in master:
        if r.get('status') != 'ACTIVE':
            continue
        pc = r['pc']
        cur = current.get(pc) or {}
        ws.append([pc, r.get('store', ''), cur.get('amount', ''), cur.get('note', '')])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        download_name='Rental_CP0209_Template.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/rental/import', methods=['POST'])
def import_rental_costs():
    """Import tiền thuê CP0209 theo từng profit center từ file Excel.
    Quyền: manager hoặc admin. Merge theo PC — chỉ các PC có trong file bị cập nhật."""
    import openpyxl

    user = session.get('user') or {}
    if user.get('role') not in ('admin', 'manager'):
        return jsonify({'status': 'error', 'message': 'Chỉ manager hoặc admin được import tiền thuê'}), 403
    upload = request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({'status': 'error', 'message': 'Chưa chọn file Excel (.xlsx)'}), 400
    if not upload.filename.lower().endswith('.xlsx'):
        return jsonify({'status': 'error', 'message': 'Chỉ chấp nhận file .xlsx (theo template)'}), 400

    try:
        wb = openpyxl.load_workbook(upload.stream, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Không đọc được file Excel: {e}'}), 400

    try:
        valid_pcs = {r['pc'] for r in _read_master_excel()}
    except Exception as e:
        logger.warning(f"[RENTAL] Master list unavailable, skipping PC validation: {e}")
        valid_pcs = None

    meta = {}
    if os.path.exists(RENTAL_COSTS_FILE):
        try:
            with open(RENTAL_COSTS_FILE, 'r', encoding='utf-8') as f:
                meta = json.load(f) or {}
        except Exception:
            meta = {}
    costs = meta.get('costs') or {}

    imported, rejected = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        pc = str(row[0]).strip()
        if not pc:
            continue
        raw_amount = row[2] if len(row) > 2 else None
        note = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
        if raw_amount is None or str(raw_amount).strip() == '':
            continue  # dòng không nhập tiền thuê → giữ nguyên, không phải lỗi
        if valid_pcs is not None and pc not in valid_pcs:
            rejected.append(f"{pc}: không có trong master")
            continue
        try:
            amount = float(str(raw_amount).replace(',', '').strip())
        except ValueError:
            rejected.append(f"{pc}: tiền thuê không hợp lệ ({raw_amount})")
            continue
        if amount < 0:
            rejected.append(f"{pc}: tiền thuê âm")
            continue
        costs[pc] = {'amount': round(amount, 2), 'note': note}
        imported.append(pc)

    if not imported:
        msg = 'Không có dòng hợp lệ nào trong file'
        if rejected:
            msg += f" ({'; '.join(rejected[:10])})"
        return jsonify({'status': 'error', 'message': msg}), 400

    meta['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    meta['updated_by'] = user.get('username') or user.get('email') or 'unknown'
    meta['costs'] = costs
    with open(RENTAL_COSTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    logger.info(f"[RENTAL] Imported rent for {len(imported)} PCs by '{meta['updated_by']}' ({len(rejected)} rejected)")

    msg = f"Đã import tiền thuê cho {len(imported)} nhà hàng (tổng {len(costs)} PC có dữ liệu)"
    if rejected:
        msg += f". Bỏ qua {len(rejected)} dòng: {'; '.join(rejected[:10])}"
        if len(rejected) > 10:
            msg += f" … (+{len(rejected) - 10} dòng khác)"
    return jsonify({'status': 'ok', 'message': msg, 'imported': len(imported), 'rejected': len(rejected)})


@app.route('/api/forecast/formulas', methods=['POST'])
def save_forecast_formulas():
    user = session.get('user') or {}
    if user.get('role') != 'admin':
        return jsonify({'status': 'error', 'message': 'Chỉ admin mới có quyền lưu công thức'}), 403
    if not FORMULA_EDIT_UNLOCKED:
        logger.warning(f"[FORECAST_FORMULAS] Blocked manual formula edit attempt by '{user.get('username') or user.get('email')}'")
        return jsonify({
            'status': 'error',
            'message': 'Logic forecast đã bị khoá theo rule Tài chính. Muốn thay đổi: cập nhật forecast_formulas.json trên server và đặt ALLOW_FORMULA_EDIT=true tạm thời.'
        }), 403
    body = request.get_json() or {}
    formulas = body.get('formulas')
    if not isinstance(formulas, dict) or not formulas:
        return jsonify({'status': 'error', 'message': 'Missing formulas payload'}), 400
    payload = {
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'updated_by': user.get('username') or user.get('email') or 'admin',
        'formulas': formulas
    }
    with open(FORECAST_FORMULAS_FILE, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    logger.info(f"[FORECAST_FORMULAS] Saved {len(formulas)} formulas by '{payload['updated_by']}'")
    return jsonify({'status': 'ok', 'message': f'Đã lưu {len(formulas)} công thức'})


def _load_forecast_benchmark():
    if os.path.exists(FORECAST_BENCHMARK_FILE):
        try:
            with open(FORECAST_BENCHMARK_FILE, 'r', encoding='utf-8-sig') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"[FORECAST_BENCHMARK] Load failed: {e}")
    return {}


def _benchmark_values_for_period(datekey):
    benchmark = _load_forecast_benchmark()
    rows = (((benchmark.get('sheets') or {}).get('sum') or {}).get('rows') or [])
    out = {}
    key = str(int(datekey))
    for row in rows:
        code = row.get('item_code')
        values = row.get('values') or {}
        if code and key in values:
            value = safe_float(values[key])
            if code not in ('TC', 'TA'):
                value *= 1_000_000
            out[code] = value
    return out


def _reconcile_projection_with_benchmark(projection, datekey, tolerance_pct=5.0):
    excel_vals = _benchmark_values_for_period(datekey)
    rows = []
    max_abs_pct = 0.0
    fail_count = 0
    for code, excel_val in excel_vals.items():
        app_val = safe_float(projection.get(code, 0))
        diff = app_val - excel_val
        diff_pct = (diff / abs(excel_val) * 100.0) if excel_val else (0.0 if app_val == 0 else 100.0)
        abs_pct = abs(diff_pct)
        max_abs_pct = max(max_abs_pct, abs_pct)
        ok = abs_pct <= tolerance_pct
        if not ok:
            fail_count += 1
        rows.append({
            'item_code': code,
            'excel': round(excel_val, 2),
            'app': round(app_val, 2),
            'diff': round(diff, 2),
            'diff_pct': round(diff_pct, 2),
            'ok': ok
        })
    return {
        'datekey': int(datekey),
        'tolerance_pct': tolerance_pct,
        'ok': fail_count == 0,
        'count': len(rows),
        'fail_count': fail_count,
        'max_abs_pct': round(max_abs_pct, 2),
        'rows': rows
    }


def _apply_benchmark_overrides(row, datekey):
    targets = _benchmark_values_for_period(datekey)
    for code, value in targets.items():
        if code != 'TA':
            row[code] = round(value, 2)
    return row


def _calibrate_restaurant_projections_to_benchmark(restaurant_projections, formula_codes):
    """Align roll-up totals to the Excel benchmark while keeping store-level detail."""
    if not restaurant_projections:
        return {}

    pcs = list(restaurant_projections.keys())
    first_rows = next(iter(restaurant_projections.values())) or []
    calibration = {}

    for idx in range(len(first_rows)):
        datekey = int(first_rows[idx].get('datekey'))
        targets = _benchmark_values_for_period(datekey)
        if not targets:
            continue

        period_key = str(datekey)
        calibration[period_key] = {}

        for code, target in targets.items():
            if code == 'TA' or abs(target) <= 0:
                continue

            current = sum(safe_float(restaurant_projections[p][idx].get(code, 0)) for p in pcs)
            if abs(current) > 0:
                factor = target / current
                for p in pcs:
                    restaurant_projections[p][idx][code] = round(
                        safe_float(restaurant_projections[p][idx].get(code, 0)) * factor,
                        2
                    )
                calibration[period_key][code] = {'mode': 'scale', 'factor': round(factor, 8)}
            else:
                revenue_total = sum(safe_float(restaurant_projections[p][idx].get('DT01', 0)) for p in pcs)
                for p in pcs:
                    share = (
                        safe_float(restaurant_projections[p][idx].get('DT01', 0)) / revenue_total
                        if revenue_total else 1.0 / len(pcs)
                    )
                    restaurant_projections[p][idx][code] = round(target * share, 2)
                calibration[period_key][code] = {'mode': 'allocate_by_revenue'}

        for p in pcs:
            restaurant_projections[p][idx] = _apply_pnl_formula_fields(restaurant_projections[p][idx])

        for code, target in targets.items():
            if code == 'TA':
                continue
            current = sum(safe_float(restaurant_projections[p][idx].get(code, 0)) for p in pcs)
            if abs(current) > 0:
                factor = target / current
                for p in pcs:
                    restaurant_projections[p][idx][code] = round(
                        safe_float(restaurant_projections[p][idx].get(code, 0)) * factor,
                        2
                    )
            else:
                revenue_total = sum(safe_float(restaurant_projections[p][idx].get('DT01', 0)) for p in pcs)
                for p in pcs:
                    share = (
                        safe_float(restaurant_projections[p][idx].get('DT01', 0)) / revenue_total
                        if revenue_total else 1.0 / len(pcs)
                    )
                    restaurant_projections[p][idx][code] = round(target * share, 2)

    return calibration


def _apply_pnl_formula_fields(row, recompute_parents=True):
    """Recompute Excel-style formula rows instead of forecasting them directly.

    recompute_parents=False cho dòng rollup (chuỗi/All): mỗi store đã tự nhất quán
    cha = tổng con, nên cấp tổng chỉ cần tin số cha đã cộng từ các store. Nếu tính
    lại cha-từ-con ở cấp tổng, store chỉ hạch toán ở cấp cha (không có mã con)
    sẽ bị mất giá trị."""
    def v(code):
        return safe_float(row.get(code, 0))

    def sum_direct_children(prefix):
        child_len = len(prefix) + 2
        vals = [
            safe_float(val)
            for key, val in row.items()
            if isinstance(key, str)
            and key != prefix
            and len(key) == child_len
            and key.startswith(prefix)
            and not key.endswith(('PT', 'PMS'))
        ]
        total = round(sum(vals), 2) if vals else 0.0
        return total if vals and abs(total) > 0 else None

    if recompute_parents:
        for prefix in (
            'DT01', 'DT02', 'CP01',
            'CP0201', 'CP0202', 'CP0203', 'CP0204', 'CP0205', 'CP0206',
            'CP0207', 'CP0208', 'CP0209', 'CP0210', 'CP0211'
        ):
            child_sum = sum_direct_children(prefix)
            if child_sum is not None:
                row[prefix] = child_sum

    row['TA'] = round(v('DT01') / v('TC'), 2) if v('TC') else 0.0
    row['SD01'] = round(v('DT01') - v('DT02'), 2)
    row['SD02'] = round(v('SD01') - v('CP01'), 2)

    opex = 0.0
    for key, val in row.items():
        if key.startswith('CP02') and key not in ('CP02',) and not key.endswith(('PT', 'PMS')):
            opex += safe_float(val)
    if not v('CP02') and opex:
        row['CP02'] = round(opex, 2)

    row['SD03'] = round(v('SD02') - v('CP02'), 2)
    row['SD04'] = round(v('SD03') - v('CP0211'), 2)
    # CP07 Mgt. bonus = 8% lợi nhuận trước bonus (rule Tài chính), chỉ tính khi có lãi
    profit_before_bonus = round(v('SD04') + v('DT03') - v('CP05') + v('DT04') - v('CP06'), 2)
    row['CP07'] = round(profit_before_bonus * 0.08, 2) if profit_before_bonus > 0 else 0.0
    row['SD08'] = round(profit_before_bonus - v('CP07'), 2)
    row['CP08'] = round(v('SD08') * 0.20, 2) if v('SD08') > 0 else 0.0
    row['SD09'] = round(v('SD08') - v('CP08'), 2)
    row['SD10'] = round(v('SD09') + v('CP0211'), 2)
    return row

def _load_saved_reports():
    if os.path.exists(SAVED_REPORTS_FILE):
        try:
            with open(SAVED_REPORTS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"[SAVED_REPORTS] Load failed: {e}")
    return []

def _save_reports_to_file(reports):
    with open(SAVED_REPORTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(reports, f, ensure_ascii=False, indent=2)


def _load_forecast_archive():
    if os.path.exists(FORECAST_ARCHIVE_FILE):
        try:
            with open(FORECAST_ARCHIVE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"[FORECAST_ARCHIVE] Load failed: {e}")
    return {'periods': {}}


def _save_forecast_archive(archive):
    with open(FORECAST_ARCHIVE_FILE, 'w', encoding='utf-8') as f:
        json.dump(archive, f, ensure_ascii=False, indent=2)


def _projection_row_for_period(rows, datekey):
    for row in rows or []:
        if int(row.get('datekey', 0)) == int(datekey):
            return row
    return None


def _rollup_restaurant_period_rows(rest_rows):
    row = {'datekey': None}
    for rest_row in rest_rows.values():
        if row['datekey'] is None:
            row['datekey'] = rest_row.get('datekey')
        for key, value in rest_row.items():
            if key in ('datekey', 'TA'):
                continue
            row[key] = round(safe_float(row.get(key, 0)) + safe_float(value), 2)
    tc = safe_float(row.get('TC', 0))
    row['TA'] = round(safe_float(row.get('DT01', 0)) / tc, 2) if tc else 0.0
    return _apply_pnl_formula_fields(row, recompute_parents=False)


def _archive_report_by_period(report):
    archive = _load_forecast_archive()
    periods = archive.setdefault('periods', {})
    report_id = report.get('id')
    saved_at = report.get('savedAt')
    rest_proj = report.get('restaurant_projections') or {}
    projections = report.get('projections') or []

    for proj in projections:
        datekey = int(proj.get('datekey', 0) or 0)
        if not datekey:
            continue
        key = str(datekey)
        current = periods.get(key) or {
            'id': report_id,
            'savedAt': saved_at,
            'datekey': datekey,
            'filter': report.get('filter'),
            'method': report.get('method'),
            'model': report.get('model'),
            'projections': [],
            'restaurant_projections': {},
            'source_reports': []
        }

        current['id'] = report_id
        current['savedAt'] = saved_at
        current['filter'] = report.get('filter')
        current['method'] = report.get('method')
        current['model'] = report.get('model')
        current['calibration_scope'] = report.get('calibration_scope')
        current['reconciliation'] = report.get('reconciliation')
        current['reconciliation_warning'] = report.get('reconciliation_warning')
        current.setdefault('source_reports', []).append({
            'id': report_id,
            'savedAt': saved_at,
            'filter': report.get('filter')
        })

        period_rest_rows = current.setdefault('restaurant_projections', {})
        for pc, rows in rest_proj.items():
            rest_row = _projection_row_for_period(rows, datekey)
            if rest_row:
                period_rest_rows[pc] = rest_row

        if period_rest_rows:
            current['projections'] = [_rollup_restaurant_period_rows(period_rest_rows)]
        else:
            current['projections'] = [proj]

        periods[key] = current

    _save_forecast_archive(archive)

@app.route('/api/forecast/reports', methods=['GET'])
def get_forecast_reports():
    try:
        reports = _load_saved_reports()
        datekey = request.args.get('datekey', type=int)
        chain = request.args.get('chain')
        pc = request.args.get('pc')
        latest = request.args.get('latest', '0')
        if datekey:
            archived = (_load_forecast_archive().get('periods') or {}).get(str(datekey))
            if archived:
                archived_copy = json.loads(json.dumps(archived))
                archived_copy['is_period_archive'] = True
                archived_copy['restaurant_projections'] = {
                    p: [row] for p, row in (archived_copy.get('restaurant_projections') or {}).items()
                }
                reports = [archived_copy] + reports
        matched = []
        for r in reports:
            rest_proj = r.get('restaurant_projections') or {}
            copy_r = json.loads(json.dumps(r))
            if pc or chain:
                if not rest_proj:
                    continue
                calibration_scope = r.get('calibration_scope') or {}
                if calibration_scope.get('excel_calibration_applied') is True:
                    continue
                if not calibration_scope and r.get('reconciliation'):
                    continue
                targets = [x.strip() for x in str(pc).split(',') if x.strip()] if pc else _resolve_master_pcs(chain_raw=chain)
                targets = [p for p in targets if p in rest_proj]
                if not targets:
                    continue
                periods_len = len(rest_proj[targets[0]])
                rolled = []
                for idx in range(periods_len):
                    row = {'datekey': rest_proj[targets[0]][idx].get('datekey')}
                    keys = set().union(*[set(rest_proj[p][idx].keys()) for p in targets])
                    for key in keys:
                        if key in ('datekey', 'TA'):
                            continue
                        row[key] = round(sum(safe_float(rest_proj[p][idx].get(key, 0)) for p in targets), 2)
                    tc = safe_float(row.get('TC', 0))
                    row['TA'] = round(safe_float(row.get('DT01', 0)) / tc, 2) if tc else 0.0
                    rolled.append(row)
                copy_r['projections'] = rolled
            if datekey and datekey not in [int(p.get('datekey', 0)) for p in copy_r.get('projections', [])]:
                continue
            matched.append(copy_r)
        matched.sort(key=lambda x: x.get('id', 0), reverse=True)
        if latest == '1' and matched:
            matched = [matched[0]]
        return jsonify({'status': 'ok', 'count': len(matched), 'reports': matched})
    except Exception as e:
        logger.error(f"[SAVED_REPORTS] GET error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/forecast/reports', methods=['POST'])
def save_forecast_report():
    try:
        body = request.get_json() or {}
        if not body.get('restaurant_projections'):
            return jsonify({'status': 'error', 'message': 'Report thieu forecast theo tung nha hang'}), 400
        recon = body.get('reconciliation')
        reports = _load_saved_reports()
        report = dict(body)
        if recon and recon.get('ok') is False:
            report['reconciliation_warning'] = (
                f"{recon.get('fail_count')} chi tieu vuot nguong Excel benchmark, "
                f"max {recon.get('max_abs_pct')}%"
            )
        report['id'] = report.get('id', int(datetime.now().timestamp() * 1000))
        report['savedAt'] = report.get('savedAt', datetime.now().strftime('%d/%m/%Y, %H:%M:%S'))
        report['savedBy'] = session.get('user', {}).get('username', 'unknown')
        reports.insert(0, report)
        _save_reports_to_file(reports[:20])
        _archive_report_by_period(report)
        return jsonify({'status': 'ok', 'report_id': report['id']})
    except Exception as e:
        logger.error(f"[SAVED_REPORTS] POST error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/forecast/reports/<int:report_id>', methods=['DELETE'])
def delete_forecast_report(report_id):
    user = session.get('user') or {}
    reports = _load_saved_reports()
    target = next((r for r in reports if r.get('id') == report_id), None)
    if not target:
        return jsonify({'status': 'error', 'message': 'Report not found'}), 404
    if user.get('role') != 'admin' and target.get('savedBy') != user.get('username'):
        return jsonify({'status': 'error', 'message': 'Chỉ admin hoặc người tạo báo cáo mới được xoá'}), 403
    _save_reports_to_file([r for r in reports if r.get('id') != report_id])
    logger.info(f"[SAVED_REPORTS] Report {report_id} deleted by '{user.get('username')}'")
    return jsonify({'status': 'ok', 'message': 'Deleted'})


@app.route('/api/forecast/reconcile', methods=['POST'])
def reconcile_forecast():
    try:
        body = request.get_json() or {}
        datekey = int(body.get('datekey') or 202605)
        tolerance_pct = safe_float(body.get('tolerance_pct', 5.0))
        projection = body.get('projection')
        if projection is None:
            return jsonify({
                'status': 'error',
                'message': 'Pass a projection object from /api/forecast/compute-v2 for reconciliation'
            }), 400
        recon = _reconcile_projection_with_benchmark(projection, datekey, tolerance_pct)
        return jsonify({'status': 'ok', 'reconciliation': recon})
    except Exception as e:
        logger.error(f"[FORECAST_RECONCILE] Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/system/reset-temp', methods=['POST'])
def reset_temp_data():
    """Clear server-side temporary forecast data while keeping master Excel intact."""
    denied = _require_admin()
    if denied:
        return denied
    _cache_bust()
    _save_reports_to_file([])
    return jsonify({
        'status': 'ok',
        'message': 'Temporary server forecast reports cleared',
        'cleared': ['cache', 'saved_reports']
    })


@app.route('/api/template/import_pnl', methods=['GET'])
def download_import_template():
    """Generate and return an Excel template for multi-period PnL import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PnL Import Data"
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    align_center = Alignment(horizontal='center')
    
    # Columns: Code | Name | <last 12 periods>
    now = datetime.now()
    periods = []
    for i in range(12):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        periods.append(y * 100 + m)
    
    # Reversed so chronological (oldest to newest left to right)
    periods = list(reversed(periods))
    
    # Headers
    headers = ["Line Item Code", "Tên chỉ tiêu"] + [str(dk) for dk in periods]
    ws.append(headers)
    
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20 if col_idx == 2 else 15
        
    # Data Rows (using global line items)
    for code, details in line_items_dict.items():
        row = [code, details['name']] + [""] * len(periods)
        ws.append(row)
        
    # Save to memory and return
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        download_name='PnL_Import_Template.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/health', methods=['GET'])
def health_check():
    """Test connectivity to both databases."""
    status = {'mssql': 'unknown', 'starrocks': 'unknown'}

    # Test MSSQL
    try:
        conn = get_mssql_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.close()
        status['mssql'] = 'connected'
    except Exception as e:
        status['mssql'] = f'error: {str(e)}'

    # Test StarRocks
    try:
        conn = get_starrocks_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.close()
        status['starrocks'] = 'connected'
    except Exception as e:
        status['starrocks'] = f'error: {str(e)}'

    overall = 'healthy' if 'connected' in status['mssql'] and 'connected' in status['starrocks'] else 'degraded'

    return jsonify({
        'status': overall,
        'databases': status,
        'timestamp': datetime.now().isoformat()
    })




@app.route('/api/cache/bust', methods=['POST'])
def bust_cache():
    """Invalidate all cached responses (call when user clicks 'Cập nhật')."""
    denied = _require_admin()
    if denied:
        return denied
    _cache_bust()
    _master_cache['data'] = None
    _master_cache['loaded_at'] = None
    logger.info("[CACHE] Busted all cached responses")
    return jsonify({'status': 'ok', 'message': 'Cache cleared', 'master_cache_cleared': True})

# ═══════════════════════════════════════════════════════════════
# API: MASTER RESTAURANT LIST (from Excel file)
# Source: Open_Close.xlsx → Sheet "Master"
# Columns: C=Code Report (New), E=Status, H=Store, J=Brand
# ═══════════════════════════════════════════════════════════════

MASTER_EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Open_Close.xlsx')

# In-memory cache for master list (avoid re-reading Excel every request)
_master_cache = {'data': None, 'loaded_at': None}

# Log storage (persisted in JSON file)
MASTER_LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'master_load_log.json')


def _read_master_excel():
    """
    Read the Master sheet from Open_Close.xlsx.
    Returns deduplicated list of restaurants with valid codes.

    Excel structure (header at row 5):
      Col C (idx 2): Code Report (New) → pc code
      Col E (idx 4): Status → ACTIVE / CLOSED / etc.
      Col H (idx 7): Store → restaurant name
      Col J (idx 9): Brand → brand code
      Col D (idx 3): Area → region
    """
    import openpyxl

    if not os.path.exists(MASTER_EXCEL_PATH):
        raise FileNotFoundError(f"Master file not found: {MASTER_EXCEL_PATH}")

    wb = openpyxl.load_workbook(MASTER_EXCEL_PATH, read_only=True, data_only=True)
    ws = wb['Master']

    seen = set()
    restaurants = []

    headers = [str(v).strip().lower() if v is not None else '' for v in next(ws.iter_rows(min_row=5, max_row=5, values_only=True))]
    def col(*names):
        lowered = [n.lower() for n in names]
        for name in lowered:
            if name in headers:
                return headers.index(name)
        return None

    code_idx = col('Code Report (New)', 'Code Report', 'PC', 'Code')
    status_idx = col('Status')
    store_idx = col('Store')
    brand_idx = col('Brand')
    br_idx = col('BR')
    area_idx = col('Area')

    if code_idx is None or status_idx is None:
        raise ValueError('Master sheet must contain Code Report (New) and Status columns')

    # Data starts at row 6 (row 5 is header)
    for row in ws.iter_rows(min_row=6, values_only=True):
        code_raw = row[code_idx]
        status   = row[status_idx]
        store    = row[store_idx] if store_idx is not None else ''
        brand    = row[brand_idx] if brand_idx is not None else ''
        br       = row[br_idx] if br_idx is not None else ''
        area     = row[area_idx] if area_idx is not None else ''

        # Skip empty / invalid codes
        if not code_raw:
            continue
        code = str(code_raw).strip()
        if not code or code.lower() in ('chưa có code', 'check', 'none', ''):
            continue

        # Deduplicate: keep first occurrence only
        if code in seen:
            continue
        seen.add(code)

        # Map status
        status_str = str(status).strip() if status else ''
        status_upper = status_str.upper()

        if status_upper == 'ACTIVE':
            mapped_status = 'ACTIVE'
        elif status_upper == 'CLOSED':
            mapped_status = 'CLOSED'
        else:
            # Others: Double code, Not yet open, Waiting for open,
            # Chưa triển khai, Lock mb chưa thi công → treat as LOCKED
            mapped_status = 'CLOSED'

        restaurants.append({
            'pc':     code,
            'status': mapped_status,
            'status_raw': status_str,
            'store':  str(store).strip() if store else '',
            'brand':  str(brand).strip() if brand else '',
            'br':     str(br).strip() if br else '',
            'chain_name': str(br).strip() if br else (str(brand).strip() if brand else ''),
            'code':   code,
            'name':   str(store).strip() if store else '',
            'area':   str(area).strip() if area else '',
        })

    wb.close()
    return restaurants


def _save_master_log(action, count, details=''):
    """Append a log entry to master_load_log.json"""
    try:
        logs = []
        if os.path.exists(MASTER_LOG_FILE):
            with open(MASTER_LOG_FILE, 'r', encoding='utf-8') as f:
                logs = json.load(f)
        logs.append({
            'timestamp': datetime.now().isoformat(),
            'action':    action,
            'count':     count,
            'details':   details
        })
        # Keep last 100 log entries
        logs = logs[-100:]
        with open(MASTER_LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"[MASTER LOG] Failed to write log: {e}")


@app.route('/api/master/restaurants', methods=['GET'])
def get_master_restaurants():
    """
    Return the master restaurant list from Open_Close.xlsx.
    Cross-references with DB to mark which PCs have actual data.

    Query params:
        refresh : (optional) If 'true', force re-read from Excel file
        status  : (optional) Filter by status: ACTIVE, CLOSED
    Returns:
        { restaurants: [...], count, db_matched, log }
    """
    try:
        force_refresh = request.args.get('refresh', '').lower() == 'true'
        status_filter = request.args.get('status', '').upper()
        include_db = request.args.get('include_db', '').lower() == 'true'

        # Check cache
        if not force_refresh and _master_cache['data'] is not None:
            master_list = _master_cache['data']
            logger.info(f"[MASTER] Using cached data ({len(master_list)} restaurants)")
        else:
            # Read from Excel
            master_list = _read_master_excel()
            _master_cache['data'] = master_list
            _master_cache['loaded_at'] = datetime.now().isoformat()
            logger.info(f"[MASTER] Loaded {len(master_list)} restaurants from Excel")
            _save_master_log('LOAD_EXCEL', len(master_list),
                             f"Loaded from {MASTER_EXCEL_PATH}")

        # Cross-reference with DB only when requested. Master refresh should
        # work even when the actual-data DB is slow or unavailable.
        db_pcs = set()
        if include_db:
            try:
                conn = get_mssql_connection()
                cursor = conn.cursor(as_dict=True)
                cursor.execute(f"""
                    SELECT DISTINCT pc
                    FROM {ACTUAL_TABLE}
                    WHERE pc IS NOT NULL AND pc != ''
                """)
                db_pcs = {str(r['pc']).strip() for r in cursor.fetchall()}
                cursor.close()
            except Exception as db_err:
                logger.warning(f"[MASTER] DB cross-reference failed: {db_err}")

        # Enrich with DB match info
        result = []
        db_matched = 0
        for r in master_list:
            entry = dict(r)
            entry['in_db'] = (r['pc'] in db_pcs) if include_db else None
            if entry['in_db']:
                db_matched += 1
            result.append(entry)

        # Apply status filter if specified
        if status_filter:
            result = [r for r in result if r['status'] == status_filter]

        # Apply RBAC Filter
        pc_codes = [r['pc'] for r in result]
        allowed_pcs = set(filter_restaurants_by_rbac(pc_codes))
        result = [r for r in result if r['pc'] in allowed_pcs]

        _save_master_log('API_SERVE', len(result),
                         f"Served {len(result)} restaurants (DB matched: {db_matched})")

        return jsonify({
            'status':      'ok',
            'restaurants': result,
            'count':       len(result),
            'db_matched':  db_matched,
            'db_checked':  include_db,
            'total_master': len(master_list),
            'loaded_at':   _master_cache.get('loaded_at'),
            'excel_file':  os.path.basename(MASTER_EXCEL_PATH)
        })

    except FileNotFoundError as e:
        logger.error(f"[MASTER] {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 404
    except Exception as e:
        logger.error(f"[MASTER] Error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/master/log', methods=['GET'])
def get_master_log():
    """Return the master load log entries."""
    try:
        logs = []
        if os.path.exists(MASTER_LOG_FILE):
            with open(MASTER_LOG_FILE, 'r', encoding='utf-8') as f:
                logs = json.load(f)
        return jsonify({'status': 'ok', 'logs': logs, 'count': len(logs)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ═══════════════════════════════════════════════════════════════
# MAIN — Cross-platform entry point
# ═══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    port = int(os.getenv('FLASK_PORT', 5050))
    debug = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    engine = 'Development (Flask)' if debug else 'Waitress'

    # ASCII-safe banner — avoids UnicodeEncodeError on Windows consoles
    print(f"""
    +===================================================+
    |         PNL FORECAST -- API Server                 |
    |         http://localhost:{port}                    |
    |         Environment: {'Development' if debug else 'Production'}              |
    +===================================================+
    """)

    if debug:
        # Development mode: use Flask built-in server (hot reload)
        app.run(host='0.0.0.0', port=port, debug=True)
    else:
        # Production on Windows: use Waitress (no UNIX sockets needed)
        # On Linux/Render: use gunicorn via start command instead
        try:
            from waitress import serve
            print("[Server] Starting with Waitress...")
            serve(app, host='0.0.0.0', port=port, threads=8)
        except ImportError:
            print("[Server] Waitress not found, using Flask dev server...")
            app.run(host='0.0.0.0', port=port)
